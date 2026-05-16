import os
import io
import json
import time
import traceback
import datetime
import cx_Oracle
from flask import Blueprint, jsonify, send_file, request, session
from sqlalchemy import text

from core.config import (
    DASHBOARD_CACHE,
    SCHEDULE_HOURS,
    ATTENTION_CACHE,
    ATTENTION_SCHEDULE_HOURS,
    QUOTA_LIMIT,
    ORACLE_CONFIG,
    ORACLE_INSTANT_CLIENT_PATH,
)
from core.database import engine
from modules.auth import admin_required, login_required, permission_required, user_has_permission, get_current_user
from modules.card_hygiene import get_all_active_jobs
from modules.contact_fallbacks import delivery_fallback_cte

dashboard_bp = Blueprint('dashboard', __name__)
QUOTA_SUMMARY_CACHE = {}
QUOTA_DAILY_CACHE = {}
QUOTA_SUMMARY_TTL_SECONDS = 120
QUOTA_DAILY_TTL_SECONDS = 120
QUOTA_CACHE_SCHEMA_VERSION = 5
QUOTA_APPLICATIONS = {
    400: 'Vale Transporte',
    500: 'Comum',
    505: 'P Social',
    910: 'Escolar',
}


def user_can_view_hygiene_dashboard():
    return user_has_permission('higienizacao')


def require_hygiene_dashboard_permission():
    if not user_can_view_hygiene_dashboard():
        return jsonify({'error': 'Acesso negado - permissao de higienizacao requerida.'}), 403
    return None


def sanitize_attention_user(user):
    if 'missing_email' in user or 'missing_phone' in user:
        missing_email = bool(user.get('missing_email'))
        missing_phone = bool(user.get('missing_phone'))
    else:
        missing_email = not bool(str(user.get('email') or '').strip())
        missing_phone = not bool(str(user.get('telefone') or '').strip())

    return {
        'tipo': user.get('tipo') or '',
        'registrado_por': user.get('registrado_por') or 'Desconhecido',
        'registro_cad': user.get('registro_cad') or '',
        'data_criacao': user.get('data_criacao') or '',
        'missing_email': missing_email,
        'missing_phone': missing_phone,
    }


def sanitize_attention_cache(cache):
    users = [sanitize_attention_user(user) for user in cache.get('users', [])]
    return {
        'last_updated': cache.get('last_updated', ''),
        'total': int(cache.get('total') or len(users)),
        'por_responsavel': cache.get('por_responsavel', {}),
        'users': users,
    }


def protected_label(value='Protegido'):
    return value


def parse_hygiene_log_clients(raw_clients):
    if not raw_clients:
        return []
    if isinstance(raw_clients, list):
        return raw_clients
    if isinstance(raw_clients, str):
        try:
            data = json.loads(raw_clients)
            return data if isinstance(data, list) else []
        except Exception:
            return []
    return []


HYGIENE_HISTORY_FIELDS = {
    'created_at': {'column': 'created_at', 'type': 'date'},
    'username': {'column': 'username', 'type': 'string'},
    'adminpanel_username': {'column': 'adminpanel_username', 'type': 'string'},
    'observation': {'column': 'observation', 'type': 'string'},
    'total_success': {'column': 'total_success', 'type': 'number'},
    'client_name': {'column': 'clients_json', 'type': 'json_text'},
    'client_cpf': {'column': 'clients_json', 'type': 'json_digits'},
    'client_card': {'column': 'clients_json', 'type': 'json_digits'},
}


def build_hygiene_history_filter_clause(filters, params):
    if not isinstance(filters, dict):
        return ''
    rules = filters.get('rules') or []
    clauses = []
    for index, rule in enumerate(rules[:12]):
        if not isinstance(rule, dict):
            continue
        field = HYGIENE_HISTORY_FIELDS.get(rule.get('field'))
        if not field:
            continue
        operator = rule.get('operator')
        value = rule.get('value')
        value_to = rule.get('value_to')
        column = field['column']
        name = f"hh_{index}"

        if field['type'] == 'date':
            if operator == 'on' and value:
                params[f'{name}_ini'] = f"{value} 00:00:00"
                params[f'{name}_fim'] = f"{value} 23:59:59"
                clauses.append(f"({column} BETWEEN :{name}_ini AND :{name}_fim)")
            elif operator == 'after' and value:
                params[name] = f"{value} 00:00:00"
                clauses.append(f"{column} >= :{name}")
            elif operator == 'before' and value:
                params[name] = f"{value} 23:59:59"
                clauses.append(f"{column} <= :{name}")
            elif operator == 'between' and value and value_to:
                params[f'{name}_ini'] = f"{value} 00:00:00"
                params[f'{name}_fim'] = f"{value_to} 23:59:59"
                clauses.append(f"({column} BETWEEN :{name}_ini AND :{name}_fim)")
            continue

        if field['type'] == 'number':
            try:
                numeric_value = int(value)
            except Exception:
                continue
            params[name] = numeric_value
            if operator == 'gte':
                clauses.append(f"{column} >= :{name}")
            elif operator == 'lte':
                clauses.append(f"{column} <= :{name}")
            elif operator == 'equals':
                clauses.append(f"{column} = :{name}")
            continue

        if value in (None, ''):
            continue
        if field['type'] == 'json_text':
            params[name] = f"%{value}%"
            clauses.append(f"COALESCE({column}, '') LIKE :{name}")
            continue
        if field['type'] == 'json_digits':
            digits = ''.join(ch for ch in str(value) if ch.isdigit())
            if digits:
                params[name] = f"%{digits}%"
                normalized_column = (
                    f"REPLACE(REPLACE(REPLACE(REPLACE(COALESCE({column}, ''), '.', ''), '-', ''), '/', ''), ' ', '')"
                )
                clauses.append(f"{normalized_column} LIKE :{name}")
            else:
                params[name] = f"%{value}%"
                clauses.append(f"COALESCE({column}, '') LIKE :{name}")
            continue
        if operator == 'equals':
            params[name] = str(value)
            clauses.append(f"{column} = :{name}")
        elif operator == 'contains':
            params[name] = f"%{value}%"
            clauses.append(f"{column} LIKE :{name}")

    return ' AND '.join(clauses)


def sanitize_hygiene_history_clients(raw_clients):
    clients = parse_hygiene_log_clients(raw_clients)
    sanitized = []
    for client in clients:
        if not isinstance(client, dict):
            continue
        cpf_digits = ''.join(ch for ch in str(client.get('cpf') or '') if ch.isdigit()).zfill(11)[-11:]
        sanitized.append({
            'cpf': cpf_digits,
            'nome': str(client.get('nome') or client.get('name') or '').strip(),
            'cartao': str(client.get('cartao') or client.get('card') or '').strip(),
            'note': str(client.get('note') or '').strip(),
            'birthdate': str(client.get('birthdate') or '').strip(),
            'birthdate_source': str(client.get('birthdate_source') or '').strip(),
        })
    return sanitized


def parse_hygiene_job_result(raw_result):
    if not raw_result:
        return {}
    if isinstance(raw_result, dict):
        return raw_result
    if isinstance(raw_result, str):
        try:
            data = json.loads(raw_result)
            return data if isinstance(data, dict) else {}
        except Exception:
            return {}
    return {}


def find_hygiene_job_result_for_log(conn, log_row):
    if not log_row:
        return {}

    created_at = log_row.get('created_at')
    params = {
        'user_id': log_row.get('user_id'),
        'username': log_row.get('username') or '',
        'observation': log_row.get('observation') or '',
        'total_success': int(log_row.get('total_success') or 0),
        'created_at': created_at,
    }

    row = conn.execute(text("""
        SELECT result_json
        FROM databridge_web.databridge_card_hygiene_jobs
        WHERE user_id = :user_id
          AND username = :username
          AND COALESCE(observation, '') = :observation
          AND CAST(JSON_UNQUOTE(JSON_EXTRACT(result_json, '$.success_count')) AS UNSIGNED) = :total_success
          AND (:created_at IS NULL OR ABS(TIMESTAMPDIFF(SECOND, COALESCE(finished_at, started_at), :created_at)) <= 86400)
        ORDER BY ABS(TIMESTAMPDIFF(SECOND, COALESCE(finished_at, started_at), :created_at)) ASC
        LIMIT 1
    """), params).mappings().first()

    return parse_hygiene_job_result(row.get('result_json')) if row else {}


def summarize_hygiene_filter(raw_filter):
    if not raw_filter:
        return 'Filtro não registrado'
    try:
        data = json.loads(raw_filter) if isinstance(raw_filter, str) else raw_filter
    except Exception:
        return 'Filtro registrado'
    labels = {
        'cpf': 'CPF',
        'cartao': 'Cartão',
        'tipo_cartao': 'Tipo',
        'app_id': 'Aplicação',
        'saldo': 'Saldo',
        'ultimo_uso': 'Último uso',
        'ultima_recarga': 'Última recarga',
        'recarga_pendente': 'Recarga pendente',
        'aluno_sem_direito': 'Aluno sem direito',
    }
    operators = {
        'equals': 'igual a',
        'contains': 'contém',
        'gte': 'maior/igual',
        'lte': 'menor/igual',
        'gt': 'maior que',
        'lt': 'menor que',
        'on': 'na data',
        'between': 'entre',
        'is_true': 'sim',
    }
    parts = []
    for rule in (data.get('rules') or [])[:4]:
        if not isinstance(rule, dict) or 'rules' in rule:
            continue
        field = labels.get(rule.get('field'), rule.get('field') or 'Campo')
        op = operators.get(rule.get('operator'), rule.get('operator') or '')
        value = rule.get('value')
        if rule.get('operator') == 'between':
            value = f"{rule.get('value')} e {rule.get('value_to')}"
        if value in (None, ''):
            value = ''
        parts.append(' '.join(str(part) for part in [field, op, value] if str(part or '').strip()))
    if not parts:
        return 'Filtro registrado'
    suffix = '...' if len(data.get('rules') or []) > 4 else ''
    return '; '.join(parts) + suffix


def build_hygiene_dashboard_payload():
    with engine.connect() as conn:
        summary_row = conn.execute(text("""
            SELECT
                COUNT(*) AS total_lotes,
                COALESCE(SUM(total_success), 0) AS total_cartoes,
                COUNT(DISTINCT user_id) AS total_operadores,
                COALESCE(SUM(CASE WHEN DATE(created_at) = CURDATE() THEN total_success ELSE 0 END), 0) AS cartoes_hoje,
                COALESCE(SUM(CASE WHEN created_at >= DATE_SUB(NOW(), INTERVAL 7 DAY) THEN total_success ELSE 0 END), 0) AS cartoes_7d
            FROM databridge_web.databridge_card_hygiene_logs
        """)).mappings().fetchone()

        ranking_rows = conn.execute(text("""
            SELECT
                user_id,
                username,
                COUNT(*) AS total_lotes,
                COALESCE(SUM(total_success), 0) AS total_cartoes,
                MAX(created_at) AS ultima_execucao
            FROM databridge_web.databridge_card_hygiene_logs
            GROUP BY user_id, username
            ORDER BY total_cartoes DESC, total_lotes DESC, username ASC
            LIMIT 15
        """)).mappings().fetchall()

        daily_rows = conn.execute(text("""
            SELECT
                DATE(created_at) AS ref_date,
                COUNT(*) AS total_lotes,
                COALESCE(SUM(total_success), 0) AS total_cartoes
            FROM databridge_web.databridge_card_hygiene_logs
            WHERE created_at >= DATE_SUB(CURDATE(), INTERVAL 13 DAY)
            GROUP BY DATE(created_at)
            ORDER BY ref_date ASC
        """)).mappings().fetchall()

        detail_rows = conn.execute(text("""
            SELECT
                id,
                user_id,
                username,
                observation,
                clients_json,
                total_success,
                created_at
            FROM databridge_web.databridge_card_hygiene_logs
            ORDER BY created_at DESC
            LIMIT 200
        """)).mappings().fetchall()

    summary = dict(summary_row or {})
    ranking = []
    for row in ranking_rows:
        item = dict(row)
        if item.get('ultima_execucao') is not None and not isinstance(item['ultima_execucao'], str):
            item['ultima_execucao'] = str(item['ultima_execucao'])
        item['total_lotes'] = int(item.get('total_lotes') or 0)
        item['total_cartoes'] = int(item.get('total_cartoes') or 0)
        ranking.append(item)

    daily_map = {}
    for row in daily_rows:
        ref_date = row.get('ref_date')
        if isinstance(ref_date, (datetime.datetime, datetime.date)):
            key = ref_date.strftime('%Y-%m-%d')
        else:
            key = str(ref_date or '')
        daily_map[key] = {
            'date': key,
            'total_lotes': int(row.get('total_lotes') or 0),
            'total_cartoes': int(row.get('total_cartoes') or 0),
        }

    daily = []
    today = datetime.date.today()
    for offset in range(13, -1, -1):
        ref_day = today - datetime.timedelta(days=offset)
        key = ref_day.strftime('%Y-%m-%d')
        daily.append(daily_map.get(key, {
            'date': key,
            'total_lotes': 0,
            'total_cartoes': 0,
        }))

    details = []
    for row in detail_rows:
        clients = parse_hygiene_log_clients(row.get('clients_json'))
        sample_cards = []
        sample_cpfs = []
        for client in clients[:3]:
            card_number = str(client.get('cartao') or '').strip()
            cpf = str(client.get('cpf') or '').strip()
            if card_number:
                sample_cards.append(card_number)
            if cpf:
                sample_cpfs.append(cpf)

        created_at = row.get('created_at')
        if created_at is not None and not isinstance(created_at, str):
            created_at = str(created_at)

        details.append({
            'id': int(row.get('id') or 0),
            'user_id': int(row.get('user_id') or 0),
            'username': row.get('username') or 'Desconhecido',
            'observation': row.get('observation') or '',
            'total_cartoes': int(row.get('total_success') or 0),
            'created_at': created_at or '',
            'sample_cards': [],
            'sample_cpfs': [],
        })

    return {
        'last_updated': datetime.datetime.now().strftime('%d/%m/%Y %H:%M'),
        'totals': {
            'total_lotes': int(summary.get('total_lotes') or 0),
            'total_cartoes': int(summary.get('total_cartoes') or 0),
            'total_operadores': int(summary.get('total_operadores') or 0),
            'cartoes_hoje': int(summary.get('cartoes_hoje') or 0),
            'cartoes_7d': int(summary.get('cartoes_7d') or 0),
        },
        'ranking': ranking,
        'daily': daily,
        'details': details,
    }


def refresh_dashboard_cache():
    """Query all 6 bases and save stats to JSON cache."""
    print(f'  [CACHE] Atualizando cache do dashboard... ({datetime.datetime.now().strftime("%H:%M:%S")})')
    stats = {}
    try:
        with engine.connect() as conn:
            stats['cad_unico'] = {
                'total': conn.execute(text("SELECT COUNT(*) FROM sntr_interligar.SALES_CAD_UNICO_JSON")).scalar(),
                'sem_email': conn.execute(text("SELECT COUNT(*) FROM sntr_interligar.SALES_CAD_UNICO_JSON WHERE email IS NULL OR email = ''")).scalar(),
                'sem_celular': conn.execute(text("SELECT COUNT(*) FROM sntr_interligar.SALES_CAD_UNICO_JSON WHERE telefone IS NULL OR telefone = ''")).scalar(),
                'sem_endereco': conn.execute(text("SELECT COUNT(*) FROM sntr_interligar.SALES_CAD_UNICO_JSON WHERE endereco IS NULL OR endereco = ''")).scalar()
            }
            q_cliente = text(
                delivery_fallback_cte()
                + """
                SELECT
                    COUNT(DISTINCT c.id) as total,
                    SUM(CASE WHEN c.email IS NULL OR c.email = '' THEN 1 ELSE 0 END) as sem_email,
                    SUM(
                        CASE
                            WHEN COALESCE(NULLIF(dl.celular_entrega, ''), NULLIF(c.cellphone, '')) IS NULL THEN 1
                            ELSE 0
                        END
                    ) as sem_celular,
                    SUM(
                        CASE
                            WHEN COALESCE(NULLIF(addr.endereco, ''), NULLIF(dl.endereco_entrega, '')) IS NULL THEN 1
                            ELSE 0
                        END
                    ) as sem_endereco
                FROM sntr_cliente.customer c
                LEFT JOIN (
                    SELECT
                        a.id_customer,
                        GROUP_CONCAT(DISTINCT CONCAT_WS(', ', a.street, a.number, a.district, a.zip_code) SEPARATOR ' | ') AS endereco
                    FROM sntr_cliente.address a
                    GROUP BY a.id_customer
                ) addr ON addr.id_customer = c.id
                LEFT JOIN delivery_latest dl
                    ON dl.cpf_limpo = REPLACE(REPLACE(c.cpf, '.', ''), '-', '')
                """
            )
            r_cli = conn.execute(q_cliente).fetchone()
            stats['clientes'] = {
                'total': r_cli[0] or 0,
                'sem_email': r_cli[1] or 0,
                'sem_celular': r_cli[2] or 0,
                'sem_endereco': r_cli[3] or 0
            }
            q_estudante = text("""
                SELECT 
                    COUNT(DISTINCT cpf) as total,
                    SUM(CASE WHEN email IS NULL OR email = '' THEN 1 ELSE 0 END) as sem_email,
                    SUM(CASE WHEN celular IS NULL OR celular = '' THEN 1 ELSE 0 END) as sem_celular,
                    SUM(CASE WHEN CONCAT_WS(', ', logradouro, numero, complemento, bairro, endereco, cep) = '' OR logradouro IS NULL THEN 1 ELSE 0 END) as sem_endereco
                FROM databridge_db.alunos
            """)
            r_est = conn.execute(q_estudante).fetchone()
            stats['estudantes'] = {
                'total': r_est[0] or 0,
                'sem_email': int(r_est[1] or 0),
                'sem_celular': int(r_est[2] or 0),
                'sem_endereco': int(r_est[3] or 0)
            }
            q_abt = text("""
                SELECT 
                    COUNT(DISTINCT documento) as total, 
                    SUM(CASE WHEN email IS NULL OR email = '' THEN 1 ELSE 0 END) as sem_email,
                    SUM(CASE WHEN celular IS NULL OR celular = '' THEN 1 ELSE 0 END) as sem_celular 
                FROM sntr_interligar.COM_CLIENTES_ABT
            """)
            r_abt = conn.execute(q_abt).fetchone()
            stats['abt'] = {
                'total': r_abt[0] or 0,
                'sem_email': int(r_abt[1] or 0),
                'sem_celular': int(r_abt[2] or 0),
                'sem_endereco': r_abt[0] or 0
            }
            q_wifi = text("SELECT COUNT(DISTINCT CPF) as total, SUM(CASE WHEN EMAIL IS NULL OR EMAIL = '' THEN 1 ELSE 0 END) as sem_email, SUM(CASE WHEN TELEFONE IS NULL OR TELEFONE = '' THEN 1 ELSE 0 END) as sem_celular FROM sntr_interligar.WIFIMAX_USERS")
            r_wifi = conn.execute(q_wifi).fetchone()
            stats['wifi'] = {
                'total': r_wifi[0] or 0,
                'sem_email': int(r_wifi[1] or 0),
                'sem_celular': int(r_wifi[2] or 0),
                'sem_endereco': r_wifi[0] or 0
            }
            total_whatsapp = conn.execute(text("SELECT COUNT(DISTINCT cpf) FROM sntr_interligar.CLIENTES_WHATSAPP")).scalar()
            stats['whatsapp'] = {
                'total': total_whatsapp or 0,
                'sem_email': total_whatsapp or 0,
                'sem_celular': 0,
                'sem_endereco': total_whatsapp or 0
            }

        for base_name, base_stats in stats.items():
            if isinstance(base_stats, dict):
                for key, value in base_stats.items():
                    if value is not None and not isinstance(value, (int, float, str, bool)):
                        stats[base_name][key] = int(value)

        cache = {
            'last_updated': datetime.datetime.now().strftime('%d/%m/%Y %H:%M'),
            'stats': stats
        }
        with open(DASHBOARD_CACHE, 'w', encoding='utf-8') as f:
            json.dump(cache, f, ensure_ascii=False)
        print('  [CACHE] Cache atualizado com sucesso.')
    except Exception as e:
        print(f'  [CACHE] Erro ao atualizar cache: {e}')


def build_attention_payload(include_sensitive=False):
    with engine.connect() as conn:
        q = text("""
            SELECT cpf, nome, tipo, data_nascimento, email, telefone, endereco, registros_cad
            FROM sntr_interligar.SALES_CAD_UNICO_JSON scuj
            WHERE
              CAST(scuj.registros_cad->>'$.data_criacao' AS DATETIME) >= DATE(NOW() - INTERVAL 7 DAY)
              AND DATE(CAST(scuj.registros_cad->>'$.data_criacao' AS DATETIME)) = DATE(CAST(scuj.registros_cad->>'$.data_registro' AS DATETIME))
              AND (
                  (email IS NULL OR TRIM(email) = '')
                  OR
                  (telefone IS NULL OR TRIM(telefone) = '')
              )
        """)
        rows = conn.execute(q).fetchall()

    users = []
    for row in rows:
        reg_cad_raw = row[7]
        reg_cad = {}
        if reg_cad_raw:
            if isinstance(reg_cad_raw, dict):
                reg_cad = reg_cad_raw
            elif isinstance(reg_cad_raw, str):
                try:
                    reg_cad = json.loads(reg_cad_raw)
                except Exception:
                    reg_cad = {}

        user = {
            'tipo': str(row[2]) if row[2] else '',
            'registrado_por': reg_cad.get('registrado_por', 'Desconhecido') or 'Desconhecido',
            'data_criacao': reg_cad.get('data_criacao', '') or '',
            'registro_cad': reg_cad.get('registro_cad', '') or reg_cad.get('criado_por', '') or 'Nao informado',
            'missing_email': not bool(str(row[4] or '').strip()),
            'missing_phone': not bool(str(row[5] or '').strip()),
        }
        if include_sensitive:
            user.update({
                'cpf': str(row[0]) if row[0] else '',
                'nome': str(row[1]) if row[1] else '',
                'data_nascimento': str(row[3]) if row[3] else '',
                'email': str(row[4]) if row[4] else '',
                'telefone': str(row[5]) if row[5] else '',
                'endereco': str(row[6]) if row[6] else '',
            })
        users.append(user)

    responsaveis = {}
    for user in users:
        resp = user['registrado_por']
        responsaveis[resp] = responsaveis.get(resp, 0) + 1

    return {
        'last_updated': datetime.datetime.now().strftime('%d/%m/%Y %H:%M'),
        'total': len(users),
        'por_responsavel': responsaveis,
        'users': users,
        'sensitive': bool(include_sensitive),
    }


def refresh_attention_cache():
    """Query incomplete registrations from last 7 days and save to JSON cache."""
    print(f'  [ATTENTION] Atualizando cache de cadastros incompletos... ({datetime.datetime.now().strftime("%H:%M:%S")})')
    try:
        cache = build_attention_payload(include_sensitive=False)
        with open(ATTENTION_CACHE, 'w', encoding='utf-8') as f:
            json.dump(cache, f, ensure_ascii=False)
        print(f"  [ATTENTION] Cache atualizado: {cache.get('total', 0)} cadastros incompletos encontrados.")
    except Exception as e:
        print(f'  [ATTENTION] Erro ao atualizar cache: {e}')


def setup_oracle_client():
    """Initializes Oracle Instant Client once."""
    try:
        cx_Oracle.init_oracle_client(lib_dir=ORACLE_INSTANT_CLIENT_PATH)
    except cx_Oracle.ProgrammingError as exc:
        if 'already been initialized' not in str(exc):
            raise


def get_quota_connection():
    setup_oracle_client()
    dsn = cx_Oracle.makedsn(
        ORACLE_CONFIG['host'],
        ORACLE_CONFIG['port'],
        service_name=ORACLE_CONFIG['database']
    )
    return cx_Oracle.connect(
        user=ORACLE_CONFIG['usuario'],
        password=ORACLE_CONFIG['senha'],
        dsn=dsn
    )


def get_quota_available_periods():
    """Returns a rolling 6-month window for quota filtering."""
    now = datetime.datetime.now()
    periods = []
    for offset in range(5, -1, -1):
        month = now.month - offset
        year = now.year
        while month <= 0:
            month += 12
            year -= 1
        periods.append((month, year))
    return periods


def serialize_quota_value(value):
    if isinstance(value, datetime.datetime):
        return value.isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, float):
        return round(value, 2)
    if isinstance(value, int):
        return value
    return value


def build_quota_transfer_exists(alias='c2'):
    return f"""
        EXISTS (
            SELECT 1
            FROM CARDACCOUNT ca_transfer
            WHERE ca_transfer.APP_ID = {alias}.APP_ID
              AND ca_transfer.USR_ID = {alias}.USR_ID
              AND ca_transfer.CAC_TRANDATE = {alias}.CAC_TRANDATE
              AND ca_transfer.CAC_TRANVALUE = {alias}.CAC_TRANVALUE
              AND ca_transfer.CAC_TYPE = 'D'
              AND (
                    ca_transfer.CI_ID <> {alias}.CI_ID
                 OR ca_transfer.CRD_SNR <> {alias}.CRD_SNR
              )
        )
    """


def build_quota_transfer_origin_card(alias='c2'):
    return f"""
        (
            SELECT MAX(formatcard(ca_transfer.ISS_ID, ca_transfer.CD_ID, ca_transfer.CRD_SNR))
            FROM CARDACCOUNT ca_transfer
            WHERE ca_transfer.APP_ID = {alias}.APP_ID
              AND ca_transfer.USR_ID = {alias}.USR_ID
              AND ca_transfer.CAC_TRANDATE = {alias}.CAC_TRANDATE
              AND ca_transfer.CAC_TRANVALUE = {alias}.CAC_TRANVALUE
              AND ca_transfer.CAC_TYPE = 'D'
              AND (
                    ca_transfer.CI_ID <> {alias}.CI_ID
                 OR ca_transfer.CRD_SNR <> {alias}.CRD_SNR
              )
        )
    """


def parse_quota_transaction_profile(raw_profile):
    profile = (raw_profile or 'all').strip().lower()
    allowed = {'all', 'purchase_only', 'transfer_only', 'mixed_only', 'admin_only'}
    return profile if profile in allowed else 'all'


def parse_quota_application_id(raw_application_id):
    try:
        application_id = int(raw_application_id or 910)
    except (TypeError, ValueError):
        application_id = 910
    return application_id if application_id in QUOTA_APPLICATIONS else 910


def build_quota_profile_filter_sql(alias, transaction_profile):
    profile = parse_quota_transaction_profile(transaction_profile)
    if profile == 'purchase_only':
        return f"WHERE {alias}.qtd_compras_reais > 0"
    if profile == 'transfer_only':
        return f"WHERE {alias}.qtd_transferencias > 0"
    if profile == 'mixed_only':
        return f"WHERE {alias}.qtd_compras_reais > 0 AND {alias}.qtd_transferencias > 0"
    if profile == 'admin_only':
        return f"WHERE {alias}.qtd_lancamentos_adm > 0"
    return ''


def build_quota_transaction_filter_sql(alias, transaction_profile):
    profile = parse_quota_transaction_profile(transaction_profile)
    if profile == 'purchase_only':
        return f"{alias}.status_movimento_key <> 'R' AND {alias}.tipo_transacao_key <> 'transferencia'"
    if profile == 'transfer_only':
        return f"{alias}.status_movimento_key <> 'R' AND {alias}.tipo_transacao_key = 'transferencia'"
    if profile == 'admin_only':
        return f"{alias}.status_movimento_key = 'R'"
    return '1 = 1'


def matches_quota_transaction_profile(purchases, transfers, transaction_profile):
    profile = parse_quota_transaction_profile(transaction_profile)
    purchases = int(purchases or 0)
    transfers = int(transfers or 0)
    if profile == 'purchase_only':
        return purchases > 0 and transfers == 0
    if profile == 'transfer_only':
        return transfers > 0 and purchases == 0
    if profile == 'mixed_only':
        return purchases > 0 and transfers > 0
    return True


def build_quota_classification_summary(purchases, transfers, admins=0):
    purchases = int(purchases or 0)
    transfers = int(transfers or 0)
    admins = int(admins or 0)
    if admins > 0 and purchases == 0 and transfers == 0:
        return 'Somente lancamento administrativo'
    if admins > 0 and purchases == 0 and transfers > 0:
        return 'Transferencias e lanc. adm.'
    if admins > 0 and purchases > 0 and transfers == 0:
        return 'Compras e lanc. adm.'
    if admins > 0 and purchases > 0 and transfers > 0:
        return 'Compras, transferencias e lanc. adm.'
    if transfers > 0 and purchases == 0:
        return 'Somente transferencia de credito'
    if purchases > 0 and transfers > 0:
        return 'Compras e transferencias'
    return 'Somente compras'


def is_quota_admin_transaction(item):
    return (item or {}).get('status_movimento_key') == 'R'


def is_quota_transfer_transaction(item):
    if is_quota_admin_transaction(item):
        return False
    return (item or {}).get('tipo_transacao_key') == 'transferencia'


def is_quota_purchase_transaction(item):
    return not is_quota_admin_transaction(item) and not is_quota_transfer_transaction(item)


def build_quota_classified_transactions_cte(date_condition_sql, extra_conditions='', application_id=910):
    extra_conditions = extra_conditions or ''
    application_id = parse_quota_application_id(application_id)
    return f"""
        WITH base_creditos AS (
            SELECT
                u.USR_ID,
                u2.USRDOC_NUMBER AS cpf,
                u.USR_NAME,
                c2.ISS_ID,
                c2.CD_ID,
                c2.CRD_SNR,
                c2.CI_ID,
                c2.APP_ID,
                c2.CAC_SEQNBR,
                c2.CAC_KEY1,
                c2.CAC_STATUS,
                formatcard(c.ISS_ID, c.CD_ID, c.CRD_SNR) AS cartao,
                c2.CAC_TRANDATE,
                c2.CAC_TRANVALUE
            FROM USERS u
            LEFT JOIN USERDOCUMENTS u2
                ON u2.USR_ID = u.USR_ID
               AND u2.DT_ID = 6
               AND u2.USRDOC_STATUS = 'A'
            INNER JOIN CARDSXUSERS c
                ON c.USR_ID = u.USR_ID
               AND c.CRDUSR_STATUS = 'A'
            INNER JOIN CARDACCOUNT c2
                ON c2.USR_ID = c.USR_ID
               AND c2.ISS_ID = c.ISS_ID
               AND c2.CD_ID = c.CD_ID
               AND c2.CRD_SNR = c.CRD_SNR
            WHERE c2.CAC_TYPE = 'C'
              AND c2.CAC_STATUS IN ('C', 'R')
              AND c2.APP_ID = {application_id}
              AND ({date_condition_sql})
              {extra_conditions}
        ),
        transferencias AS (
            SELECT
                bc.USR_ID,
                bc.CAC_SEQNBR,
                MAX(formatcard(d.ISS_ID, d.CD_ID, d.CRD_SNR)) AS cartao_origem_transferencia
            FROM base_creditos bc
            INNER JOIN CARDACCOUNT d
                ON d.APP_ID = bc.APP_ID
               AND d.USR_ID = bc.USR_ID
               AND d.CAC_TRANDATE = bc.CAC_TRANDATE
               AND d.CAC_TRANVALUE = bc.CAC_TRANVALUE
               AND d.CAC_TYPE = 'D'
               AND (
                    d.CI_ID <> bc.CI_ID
                 OR d.CRD_SNR <> bc.CRD_SNR
               )
            WHERE bc.CAC_KEY1 = 99999
            GROUP BY bc.USR_ID, bc.CAC_SEQNBR
        ),
        movimentos_classificados AS (
            SELECT
                bc.USR_ID,
                bc.cpf,
                bc.USR_NAME,
                bc.cartao,
                bc.CAC_TRANDATE AS data_hora_compra,
                bc.CAC_TRANVALUE / 100 AS valor_transacao,
                bc.CAC_STATUS AS status_movimento_key,
                CASE
                    WHEN bc.CAC_STATUS = 'R' THEN 'Lancamento Administrativo'
                    ELSE 'Concluido'
                END AS status_movimento,
                CASE
                    WHEN bc.CAC_KEY1 = 99999 OR tf.CAC_SEQNBR IS NOT NULL THEN 'transferencia'
                    ELSE 'compra'
                END AS tipo_transacao_key,
                CASE
                    WHEN bc.CAC_KEY1 = 99999 OR tf.CAC_SEQNBR IS NOT NULL THEN 'Transferencia de Credito'
                    ELSE 'Compra'
                END AS tipo_transacao,
                tf.cartao_origem_transferencia,
                TO_CHAR(bc.CAC_TRANDATE, 'MM-YYYY') AS periodo
            FROM base_creditos bc
            LEFT JOIN transferencias tf
                ON tf.USR_ID = bc.USR_ID
               AND tf.CAC_SEQNBR = bc.CAC_SEQNBR
        )
    """


def build_quota_summary_transactions_cte(date_condition_sql, extra_conditions='', application_id=910):
    extra_conditions = extra_conditions or ''
    application_id = parse_quota_application_id(application_id)
    return f"""
        WITH movimentos_classificados AS (
            SELECT
                u.USR_ID,
                u2.USRDOC_NUMBER AS cpf,
                u.USR_NAME,
                formatcard(c2.ISS_ID, c2.CD_ID, c2.CRD_SNR) AS cartao,
                c2.CAC_TRANDATE AS data_hora_compra,
                c2.CAC_TRANVALUE / 100 AS valor_transacao,
                c2.CAC_STATUS AS status_movimento_key,
                CASE
                    WHEN c2.CAC_KEY1 = 99999 OR EXISTS (
                        SELECT 1
                        FROM CARDACCOUNT d
                        WHERE d.APP_ID = c2.APP_ID
                          AND d.USR_ID = c2.USR_ID
                          AND d.CAC_TRANDATE = c2.CAC_TRANDATE
                          AND d.CAC_TRANVALUE = c2.CAC_TRANVALUE
                          AND d.CAC_TYPE = 'D'
                          AND (
                              d.CI_ID <> c2.CI_ID
                              OR d.CRD_SNR <> c2.CRD_SNR
                          )
                    ) THEN 'transferencia'
                    ELSE 'compra'
                END AS tipo_transacao_key,
                TO_CHAR(c2.CAC_TRANDATE, 'MM-YYYY') AS periodo
            FROM USERS u
            LEFT JOIN USERDOCUMENTS u2
                ON u2.USR_ID = u.USR_ID
               AND u2.DT_ID = 6
               AND u2.USRDOC_STATUS = 'A'
            INNER JOIN CARDACCOUNT c2
                ON c2.USR_ID = u.USR_ID
            WHERE c2.CAC_TYPE = 'C'
              AND c2.CAC_STATUS IN ('C', 'R')
              AND c2.APP_ID = {application_id}
              AND ({date_condition_sql})
              {extra_conditions}
        )
    """


def fetch_quota_period_rows(cursor, month, year, quota_limit, application_id=910):
    if month == 12:
        next_month, next_year = 1, year + 1
    else:
        next_month, next_year = month + 1, year

    date_condition = (
        f"c2.CAC_TRANDATE >= TO_DATE('{month:02d}/{year}', 'MM/YYYY') "
        f"AND c2.CAC_TRANDATE < TO_DATE('{next_month:02d}/{next_year}', 'MM/YYYY')"
    )
    query = f"""
        {build_quota_classified_transactions_cte(date_condition, application_id=application_id)}
        , detalhe_compras AS (
            SELECT
                mc.USR_ID,
                mc.cpf,
                mc.USR_NAME,
                mc.cartao,
                mc.data_hora_compra,
                mc.valor_transacao,
                mc.status_movimento_key,
                mc.status_movimento,
                mc.tipo_transacao_key,
                mc.tipo_transacao,
                mc.cartao_origem_transferencia,
                mc.periodo,
                SUM(mc.valor_transacao) OVER (
                    PARTITION BY mc.USR_ID, TRUNC(mc.data_hora_compra, 'MM')
                ) AS total_acumulado_estouro
            FROM movimentos_classificados mc
        )
        SELECT
            USR_ID,
            cpf,
            USR_NAME,
            cartao,
            data_hora_compra,
            valor_transacao,
            status_movimento_key,
            status_movimento,
            tipo_transacao_key,
            tipo_transacao,
            cartao_origem_transferencia,
            periodo,
            total_acumulado_estouro
        FROM detalhe_compras
        WHERE total_acumulado_estouro > {quota_limit}
        ORDER BY total_acumulado_estouro DESC, USR_NAME, data_hora_compra
    """

    cursor.execute(query)
    columns = [col[0].lower() for col in cursor.description]
    rows = []
    for row in cursor.fetchall():
        record = {}
        for col_name, value in zip(columns, row):
            record[col_name] = serialize_quota_value(value)
        rows.append(record)
    return rows


def parse_quota_limit(raw_limit):
    try:
        limit = float(raw_limit or QUOTA_LIMIT)
    except (TypeError, ValueError):
        limit = float(QUOTA_LIMIT)
    return max(limit, 0)


def filter_quota_data_by_profile(filtered_data, transaction_profile):
    profile = parse_quota_transaction_profile(transaction_profile)
    if profile == 'all':
        return filtered_data

    if profile == 'admin_only':
        return [item for item in filtered_data if item.get('status_movimento_key') == 'R']
    if profile == 'transfer_only':
        return [item for item in filtered_data if is_quota_transfer_transaction(item)]
    if profile == 'purchase_only':
        return [item for item in filtered_data if is_quota_purchase_transaction(item)]
    return filtered_data


def fetch_quota_data(period_keys, quota_limit, transaction_profile='all', application_id=910):
    """Fetch quota data on demand for selected periods and limit."""
    conn = None
    try:
        conn = get_quota_connection()
        cursor = conn.cursor()
        data = []
        for period_key in period_keys:
            try:
                month_str, year_str = period_key.split('-')
                month = int(month_str)
                year = int(year_str)
            except ValueError:
                continue
            data.extend(fetch_quota_period_rows(cursor, month, year, quota_limit, application_id))
        return filter_quota_data_by_profile(data, transaction_profile)
    finally:
        if conn:
            conn.close()


def build_quota_date_filter(period_keys):
    clauses = []
    for period_key in period_keys:
        try:
            month_str, year_str = period_key.split('-')
            month = int(month_str)
            year = int(year_str)
        except ValueError:
            continue

        if month == 12:
            next_month, next_year = 1, year + 1
        else:
            next_month, next_year = month + 1, year

        clauses.append(
            f"(c2.CAC_TRANDATE >= TO_DATE('{month:02d}/{year}', 'MM/YYYY') "
            f"AND c2.CAC_TRANDATE < TO_DATE('{next_month:02d}/{next_year}', 'MM/YYYY'))"
        )
    return ' OR '.join(clauses) if clauses else '1 = 0'


def fetch_quota_summary_data(period_keys, quota_limit, transaction_profile='all', application_id=910):
    """Fetch summary data with one aggregated Oracle query for all selected periods."""
    conn = None
    try:
        conn = get_quota_connection()
        cursor = conn.cursor()
        date_filter = build_quota_date_filter(period_keys)
        transaction_filter_sql = build_quota_transaction_filter_sql('mc', transaction_profile)

        summary_query = f"""
            {build_quota_classified_transactions_cte(date_filter, application_id=application_id)}
            , periodos_excedentes AS (
                SELECT
                    mc.USR_ID,
                    TRUNC(mc.data_hora_compra, 'MM') AS mes_referencia
                FROM movimentos_classificados mc
                GROUP BY mc.USR_ID, TRUNC(mc.data_hora_compra, 'MM')
                HAVING SUM(mc.valor_transacao) > {quota_limit}
            ),
            periodos_usuario AS (
                SELECT
                    mc.USR_ID,
                    MAX(mc.cpf) AS cpf,
                    MAX(mc.USR_NAME) AS usr_name,
                    MAX(mc.cartao) AS cartao_principal,
                    TRUNC(mc.data_hora_compra, 'MM') AS mes_referencia,
                    SUM(mc.valor_transacao) AS total_periodo,
                    COUNT(*) AS qtd_transacoes,
                    SUM(CASE WHEN mc.status_movimento_key = 'R' THEN 1 ELSE 0 END) AS qtd_lancamentos_adm,
                    SUM(
                        CASE
                            WHEN mc.status_movimento_key = 'R' THEN 0
                            WHEN mc.tipo_transacao_key = 'transferencia' THEN 0
                            ELSE 1
                        END
                    ) AS qtd_compras_reais,
                    SUM(
                        CASE
                            WHEN mc.status_movimento_key = 'R' THEN 0
                            WHEN mc.tipo_transacao_key = 'transferencia' THEN 1
                            ELSE 0
                        END
                    ) AS qtd_transferencias
                FROM movimentos_classificados mc
                INNER JOIN periodos_excedentes pe
                    ON pe.USR_ID = mc.USR_ID
                   AND pe.mes_referencia = TRUNC(mc.data_hora_compra, 'MM')
                WHERE {transaction_filter_sql}
                GROUP BY mc.USR_ID, TRUNC(mc.data_hora_compra, 'MM')
            )
            SELECT
                usr_id,
                MAX(cpf) AS cpf,
                MAX(usr_name) AS usr_name,
                MAX(cartao_principal) AS cartao_principal,
                COUNT(DISTINCT cartao_principal) AS qtd_cartoes,
                SUM(total_periodo) AS total_comprado,
                SUM(qtd_transacoes) AS qtd_transacoes,
                SUM(qtd_lancamentos_adm) AS qtd_lancamentos_adm,
                SUM(qtd_compras_reais) AS qtd_compras_reais,
                SUM(qtd_transferencias) AS qtd_transferencias,
                COUNT(*) AS qtd_periodos,
                SUM(GREATEST(total_periodo - {quota_limit}, 0)) AS total_excedente
            FROM periodos_usuario
            GROUP BY usr_id
        """
        cursor.execute(summary_query)

        ranking = []
        total_transactions_overall = 0
        total_value_overall = 0.0
        for row in cursor.fetchall():
            qtd_compras_reais = int(row[8] or 0)
            qtd_transferencias = int(row[9] or 0)
            qtd_lancamentos_adm = int(row[7] or 0)
            total_comprado = round(float(row[5] or 0), 2)
            total_transactions = int(row[6] or 0)
            ranking.append({
                'user_id': int(row[0]),
                'cpf': str(row[1] or ''),
                'usr_name': str(row[2] or 'Nao informado'),
                'cartao_principal': str(row[3] or ''),
                'qtd_cartoes': int(row[4] or 0),
                'total_comprado': total_comprado,
                'qtd_transacoes': total_transactions,
                'qtd_lancamentos_adm': qtd_lancamentos_adm,
                'qtd_compras': qtd_compras_reais,
                'qtd_compras_reais': qtd_compras_reais,
                'qtd_transferencias': qtd_transferencias,
                'qtd_periodos': int(row[10] or 0),
                'total_excedente': round(float(row[11] or 0), 2),
                'classificacao_resumo': build_quota_classification_summary(
                    qtd_compras_reais,
                    qtd_transferencias,
                    qtd_lancamentos_adm
                ),
                'somente_transferencias': (
                    qtd_transferencias > 0 and qtd_compras_reais == 0 and qtd_lancamentos_adm == 0
                ),
            })
            total_transactions_overall += total_transactions
            total_value_overall += total_comprado

        ranking.sort(key=lambda item: (-item['total_comprado'], item['usr_name']))

        return {
            'ranking': ranking,
            'daily': [],
            'totals': {
                'users': len(ranking),
                'transactions': total_transactions_overall,
                'value': round(total_value_overall, 2),
            }
        }
    finally:
        if conn:
            conn.close()


def fetch_quota_daily_data(period_keys, quota_limit, transaction_profile='all', application_id=910):
    """Fetch chart data aggregated in Oracle instead of loading the full detail list."""
    conn = None
    try:
        conn = get_quota_connection()
        cursor = conn.cursor()
        date_filter = build_quota_date_filter(period_keys)
        transaction_filter_sql = build_quota_transaction_filter_sql('mc', transaction_profile)

        query = f"""
            {build_quota_classified_transactions_cte(date_filter, application_id=application_id)}
            , periodos_excedentes AS (
                SELECT
                    mc.USR_ID,
                    TRUNC(mc.data_hora_compra, 'MM') AS mes_referencia
                FROM movimentos_classificados mc
                GROUP BY mc.USR_ID, TRUNC(mc.data_hora_compra, 'MM')
                HAVING SUM(mc.valor_transacao) > {quota_limit}
            ),
            periodos_usuario AS (
                SELECT
                    mc.USR_ID,
                    TRUNC(mc.data_hora_compra, 'MM') AS mes_referencia,
                    SUM(mc.valor_transacao) AS total_periodo,
                    SUM(CASE WHEN mc.status_movimento_key = 'R' THEN 1 ELSE 0 END) AS qtd_lancamentos_adm,
                    SUM(
                        CASE
                            WHEN mc.status_movimento_key = 'R' THEN 0
                            WHEN mc.tipo_transacao_key = 'transferencia' THEN 0
                            ELSE 1
                        END
                    ) AS qtd_compras_reais,
                    SUM(
                        CASE
                            WHEN mc.status_movimento_key = 'R' THEN 0
                            WHEN mc.tipo_transacao_key = 'transferencia' THEN 1
                            ELSE 0
                        END
                    ) AS qtd_transferencias
                FROM movimentos_classificados mc
                INNER JOIN periodos_excedentes pe
                    ON pe.USR_ID = mc.USR_ID
                   AND pe.mes_referencia = TRUNC(mc.data_hora_compra, 'MM')
                WHERE {transaction_filter_sql}
                GROUP BY mc.USR_ID, TRUNC(mc.data_hora_compra, 'MM')
            )
            SELECT
                TO_CHAR(TRUNC(mc.data_hora_compra), 'YYYY-MM-DD') AS dia,
                ROUND(SUM(mc.valor_transacao), 2) AS total_dia
            FROM movimentos_classificados mc
            INNER JOIN periodos_usuario pf
                ON pf.usr_id = mc.usr_id
               AND pf.mes_referencia = TRUNC(mc.data_hora_compra, 'MM')
            WHERE {transaction_filter_sql}
            GROUP BY TRUNC(mc.data_hora_compra)
            ORDER BY TRUNC(mc.data_hora_compra)
        """
        cursor.execute(query)
        return [
            {'date': row[0], 'total': round(float(row[1] or 0), 2)}
            for row in cursor.fetchall()
        ]
    finally:
        if conn:
            conn.close()


def build_quota_dashboard_payload(filtered_data, quota_limit):
    user_map = {}
    daily_map = {}

    for item in filtered_data:
        user_key = str(item.get('usr_id') or item.get('cpf') or item.get('usr_name') or '')
        if not user_key:
            continue

        user = user_map.setdefault(user_key, {
            'user_id': int(item.get('usr_id') or 0),
            'cpf': str(item.get('cpf') or ''),
            'usr_name': str(item.get('usr_name') or 'Nao informado'),
            'cartoes': set(),
            'period_totals': {},
            'qtd_compras': 0,
            'qtd_compras_reais': 0,
            'qtd_transferencias': 0,
            'qtd_lancamentos_adm': 0,
        })

        if item.get('cartao'):
            user['cartoes'].add(str(item.get('cartao')))
        user['qtd_compras'] += 1
        if item.get('status_movimento_key') == 'R':
            user['qtd_lancamentos_adm'] += 1
        if is_quota_transfer_transaction(item):
            user['qtd_transferencias'] += 1
        elif is_quota_purchase_transaction(item):
            user['qtd_compras_reais'] += 1
        period = item.get('periodo') or ''
        total_period = float(item.get('total_acumulado_estouro') or 0)
        user['period_totals'][period] = max(user['period_totals'].get(period, 0), total_period)

        purchase_date = item.get('data_hora_compra')
        date_key = ''
        if isinstance(purchase_date, str) and purchase_date:
            date_key = purchase_date[:10]
        elif isinstance(purchase_date, (datetime.date, datetime.datetime)):
            date_key = purchase_date.strftime('%Y-%m-%d')

        if date_key:
            daily_map[date_key] = round(
                daily_map.get(date_key, 0) + float(item.get('valor_transacao') or 0),
                2
            )

    ranking = []
    for user in user_map.values():
        total_comprado = sum(user['period_totals'].values())
        total_excedente = sum(max(total - quota_limit, 0) for total in user['period_totals'].values())
        ranking.append({
            'user_id': user['user_id'],
            'cpf': user['cpf'],
            'usr_name': user['usr_name'],
            'cartao_principal': sorted(user['cartoes'])[0] if user['cartoes'] else '',
            'qtd_cartoes': len(user['cartoes']),
            'total_comprado': round(total_comprado, 2),
            'qtd_compras': user['qtd_compras'],
            'qtd_compras_reais': user['qtd_compras_reais'],
            'qtd_transferencias': user['qtd_transferencias'],
            'qtd_lancamentos_adm': user['qtd_lancamentos_adm'],
            'qtd_periodos': len(user['period_totals']),
            'total_excedente': round(total_excedente, 2),
            'classificacao_resumo': build_quota_classification_summary(
                user['qtd_compras_reais'],
                user['qtd_transferencias'],
                user['qtd_lancamentos_adm']
            ),
            'somente_transferencias': (
                user['qtd_transferencias'] > 0
                and user['qtd_compras_reais'] == 0
                and user['qtd_lancamentos_adm'] == 0
            ),
        })

    ranking.sort(key=lambda item: (-item['total_comprado'], item['usr_name']))
    daily = [
        {'date': date_key, 'total': round(total, 2)}
        for date_key, total in sorted(daily_map.items())
    ]

    total_transactions = sum(item['qtd_compras'] for item in ranking)
    total_value = round(sum(item['total_comprado'] for item in ranking), 2)

    return {
        'ranking': ranking,
        'daily': daily,
        'totals': {
            'users': len(ranking),
            'transactions': total_transactions,
            'value': total_value,
        }
    }


def get_quota_summary_cache_key(period_keys, quota_limit, transaction_profile='all', application_id=910):
    rounded_limit = round(float(quota_limit), 2)
    return (
        QUOTA_CACHE_SCHEMA_VERSION,
        '|'.join(sorted(period_keys)),
        rounded_limit,
        parse_quota_transaction_profile(transaction_profile),
        parse_quota_application_id(application_id)
    )


def get_cached_quota_summary(period_keys, quota_limit, transaction_profile='all', application_id=910):
    cache_key = get_quota_summary_cache_key(period_keys, quota_limit, transaction_profile, application_id)
    cached = QUOTA_SUMMARY_CACHE.get(cache_key)
    if not cached:
        return None

    now_ts = time.time()
    if now_ts - cached['ts'] > QUOTA_SUMMARY_TTL_SECONDS:
        QUOTA_SUMMARY_CACHE.pop(cache_key, None)
        return None
    return cached['payload']


def store_cached_quota_summary(period_keys, quota_limit, payload, transaction_profile='all', application_id=910):
    cache_key = get_quota_summary_cache_key(period_keys, quota_limit, transaction_profile, application_id)
    QUOTA_SUMMARY_CACHE[cache_key] = {
        'ts': time.time(),
        'payload': payload
    }


def get_cached_quota_daily(period_keys, quota_limit, transaction_profile='all', application_id=910):
    cache_key = get_quota_summary_cache_key(period_keys, quota_limit, transaction_profile, application_id)
    cached = QUOTA_DAILY_CACHE.get(cache_key)
    if not cached:
        return None

    now_ts = time.time()
    if now_ts - cached['ts'] > QUOTA_DAILY_TTL_SECONDS:
        QUOTA_DAILY_CACHE.pop(cache_key, None)
        return None
    return cached['payload']


def store_cached_quota_daily(period_keys, quota_limit, payload, transaction_profile='all', application_id=910):
    cache_key = get_quota_summary_cache_key(period_keys, quota_limit, transaction_profile, application_id)
    QUOTA_DAILY_CACHE[cache_key] = {
        'ts': time.time(),
        'payload': payload
    }


def fetch_quota_user_details(user_id, period_keys, quota_limit, application_id=910, transaction_profile='all'):
    """Fetch detailed transactions only for the selected user."""
    conn = None
    try:
        conn = get_quota_connection()
        cursor = conn.cursor()
        date_filter = build_quota_date_filter(period_keys)
        query = f"""
            {build_quota_classified_transactions_cte(date_filter, f'AND u.USR_ID = {int(user_id)}', application_id)}
            , detalhe_compras AS (
                SELECT
                    mc.USR_ID,
                    mc.cpf,
                    mc.USR_NAME,
                    mc.cartao,
                    mc.data_hora_compra,
                    mc.valor_transacao,
                    mc.status_movimento_key,
                    mc.status_movimento,
                    mc.tipo_transacao_key,
                    mc.tipo_transacao,
                    mc.cartao_origem_transferencia,
                    mc.periodo,
                    SUM(mc.valor_transacao) OVER (
                        PARTITION BY mc.USR_ID, TRUNC(mc.data_hora_compra, 'MM')
                    ) AS total_mes_estouro
                FROM movimentos_classificados mc
            )
            SELECT
                USR_ID,
                cpf,
                USR_NAME,
                cartao,
                data_hora_compra,
                valor_transacao,
                status_movimento_key,
                status_movimento,
                tipo_transacao_key,
                tipo_transacao,
                cartao_origem_transferencia,
                periodo,
                total_mes_estouro
            FROM detalhe_compras
            WHERE total_mes_estouro > {quota_limit}
            ORDER BY data_hora_compra
        """
        cursor.execute(query)
        details = []
        columns = [col[0].lower() for col in cursor.description]
        for row in cursor.fetchall():
            record = {}
            for col_name, value in zip(columns, row):
                record[col_name] = serialize_quota_value(value)
            details.append(record)
        return filter_quota_data_by_profile(details, transaction_profile)
    finally:
        if conn:
            conn.close()


def fetch_quota_user_has_more_details(user_id, period_keys, quota_limit, application_id=910, transaction_profile='all'):
    profile = parse_quota_transaction_profile(transaction_profile)
    if profile == 'all':
        return False
    all_details = fetch_quota_user_details(user_id, period_keys, quota_limit, application_id, 'all')
    filtered_details = filter_quota_data_by_profile(all_details, profile)
    return len(all_details) > len(filtered_details)


def dashboard_scheduler():
    """Background thread: refreshes dashboard and attention caches at scheduled hours."""
    refreshed_dashboard = set()
    refreshed_attention = set()

    now = datetime.datetime.now()
    today = now.strftime('%Y-%m-%d')
    current_hour = now.hour

    if os.path.exists(DASHBOARD_CACHE):
        for scheduled_hour in SCHEDULE_HOURS:
            if current_hour >= scheduled_hour:
                refreshed_dashboard.add(f'{today}-dash-{scheduled_hour}')

    if os.path.exists(ATTENTION_CACHE):
        for scheduled_hour in ATTENTION_SCHEDULE_HOURS:
            if current_hour >= scheduled_hour:
                refreshed_attention.add(f'{today}-att-{scheduled_hour}')

    while True:
        now = datetime.datetime.now()
        today = now.strftime('%Y-%m-%d')
        hour = now.hour

        for scheduled_hour in SCHEDULE_HOURS:
            key = f'{today}-dash-{scheduled_hour}'
            if hour >= scheduled_hour and key not in refreshed_dashboard:
                refreshed_dashboard.add(key)
                refresh_dashboard_cache()

        for scheduled_hour in ATTENTION_SCHEDULE_HOURS:
            key = f'{today}-att-{scheduled_hour}'
            if hour >= scheduled_hour and key not in refreshed_attention:
                refreshed_attention.add(key)
                refresh_attention_cache()

        refreshed_dashboard = {key for key in refreshed_dashboard if key.startswith(today)}
        refreshed_attention = {key for key in refreshed_attention if key.startswith(today)}
        time.sleep(60)


def load_cached_json(cache_path, refresh_fn):
    if not os.path.exists(cache_path):
        refresh_fn()
    with open(cache_path, 'r', encoding='utf-8-sig') as f:
        return json.load(f)


def parse_quota_periods_param(periods_param, available_periods):
    available_keys = {item.get('key') for item in available_periods or []}
    if not periods_param:
        now = datetime.datetime.now()
        current_key = f'{now.month:02d}-{now.year}'
        if current_key in available_keys:
            return [current_key]
        return [max(available_keys)] if available_keys else []

    selected = []
    for raw_item in periods_param.split(','):
        item = raw_item.strip()
        if item and item in available_keys:
            selected.append(item)
    return sorted(set(selected))


def build_quota_summary_rows(filtered_data, quota_limit):
    user_map = {}
    for item in filtered_data:
        user_key = str(item.get('usr_id') or item.get('cpf') or item.get('usr_name') or '')
        if not user_key:
            continue
        user = user_map.setdefault(user_key, {
            'cpf': str(item.get('cpf') or ''),
            'nome': str(item.get('usr_name') or 'Nao informado'),
            'period_totals': {},
            'transactions': 0,
            'qtd_compras_reais': 0,
            'qtd_transferencias': 0,
            'qtd_lancamentos_adm': 0,
            'cards': set(),
        })
        user['transactions'] += 1
        if item.get('status_movimento_key') == 'R':
            user['qtd_lancamentos_adm'] += 1
        if is_quota_transfer_transaction(item):
            user['qtd_transferencias'] += 1
        elif is_quota_purchase_transaction(item):
            user['qtd_compras_reais'] += 1
        if item.get('cartao'):
            user['cards'].add(str(item.get('cartao')))
        period = item.get('periodo') or ''
        total_period = float(item.get('total_acumulado_estouro') or 0)
        user['period_totals'][period] = max(user['period_totals'].get(period, 0), total_period)

    summary_rows = []
    for user in user_map.values():
        total_comprado = sum(user['period_totals'].values())
        total_excedente = sum(max(total - quota_limit, 0) for total in user['period_totals'].values())
        summary_rows.append({
            'cpf': user['cpf'],
            'nome': user['nome'],
            'cartoes': ' | '.join(sorted(user['cards'])),
            'periodos': ', '.join(sorted([p for p in user['period_totals'].keys() if p])),
            'qtd_periodos': len(user['period_totals']),
            'qtd_transacoes': user['transactions'],
            'qtd_compras': user['qtd_compras_reais'],
            'qtd_transferencias': user['qtd_transferencias'],
            'qtd_lancamentos_adm': user['qtd_lancamentos_adm'],
            'total_comprado': round(total_comprado, 2),
            'total_excedente': round(total_excedente, 2),
            'classificacao_resumo': build_quota_classification_summary(
                user['qtd_compras_reais'],
                user['qtd_transferencias'],
                user['qtd_lancamentos_adm']
            ),
        })

    summary_rows.sort(key=lambda row: (-row['total_comprado'], row['nome']))
    return summary_rows


@dashboard_bp.route('/api/dashboard_stats', methods=['GET'])
@login_required
@permission_required('dashboard')
def get_dashboard_stats():
    """Returns cached dashboard stats from JSON file."""
    try:
        cache = load_cached_json(DASHBOARD_CACHE, refresh_dashboard_cache)
        result = cache.get('stats', {})
        result['last_updated'] = cache.get('last_updated', 'Desconhecido')
        return jsonify(result), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/attention_users', methods=['GET'])
@login_required
@permission_required('dashboard')
def get_attention_users():
    """Returns cached attention panel data from JSON file."""
    try:
        return jsonify(build_attention_payload(include_sensitive=True)), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/hygiene_dashboard_stats', methods=['GET'])
@login_required
def get_hygiene_dashboard_stats():
    permission_error = require_hygiene_dashboard_permission()
    if permission_error:
        return permission_error

    try:
        payload = build_hygiene_dashboard_payload()
        return jsonify(payload), 200
    except Exception as e:
        print('  [HYGIENE_DASHBOARD] Erro em /api/hygiene_dashboard_stats:')
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/admin/hygiene_jobs', methods=['GET'])
@admin_required
def admin_hygiene_jobs():
    return jsonify({'jobs': get_all_active_jobs()})


@dashboard_bp.route('/api/hygiene_history', methods=['POST'])
@login_required
def get_hygiene_history():
    permission_error = require_hygiene_dashboard_permission()
    if permission_error:
        return permission_error

    payload = request.json or {}
    filters = payload.get('filters') or {}
    page = max(1, int(payload.get('page') or 1))
    per_page = min(100, max(10, int(payload.get('per_page') or 15)))
    offset = (page - 1) * per_page

    current_user = get_current_user()
    is_admin = bool(current_user and current_user.get('is_admin'))

    params = {}
    where = []
    if not is_admin:
        where.append('user_id = :session_user_id')
        params['session_user_id'] = session.get('user_id')

    filter_sql = build_hygiene_history_filter_clause(filters, params)
    if filter_sql:
        where.append(filter_sql)

    where_sql = ' AND '.join(where) if where else '1=1'

    try:
        with engine.connect() as conn:
            total = conn.execute(text(f"""
                SELECT COUNT(1)
                FROM databridge_web.databridge_card_hygiene_logs
                WHERE {where_sql}
            """), params).scalar() or 0

            query_params = dict(params)
            query_params.update({'limit': per_page, 'offset': offset})
            rows = conn.execute(text(f"""
                SELECT id, user_id, username, adminpanel_username, observation, filter_json,
                       total_success, created_at
                FROM databridge_web.databridge_card_hygiene_logs
                WHERE {where_sql}
                ORDER BY created_at DESC
                LIMIT :limit OFFSET :offset
            """), query_params).mappings().fetchall()

        items = []
        for row in rows:
            created_at = row.get('created_at')
            if created_at is not None and not isinstance(created_at, str):
                created_at = created_at.strftime('%d/%m/%Y %H:%M')
            items.append({
                'id': int(row.get('id') or 0),
                'username': row.get('username') or 'Desconhecido',
                'adminpanel_username': row.get('adminpanel_username') or '-',
                'observation': row.get('observation') or '',
                'filter_summary': summarize_hygiene_filter(row.get('filter_json')),
                'total_success': int(row.get('total_success') or 0),
                'created_at': created_at or '',
            })

        return jsonify({
            'items': items,
            'total': int(total),
            'page': page,
            'per_page': per_page,
            'total_pages': (int(total) + per_page - 1) // per_page,
            'scope': 'all' if is_admin else 'own',
        }), 200
    except Exception as e:
        print('  [HYGIENE_HISTORY] Erro em /api/hygiene_history:')
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/hygiene_history/<int:log_id>', methods=['GET'])
@login_required
def get_hygiene_history_detail(log_id):
    permission_error = require_hygiene_dashboard_permission()
    if permission_error:
        return permission_error

    current_user = get_current_user()
    is_admin = bool(current_user and current_user.get('is_admin'))
    params = {'log_id': log_id}
    scope_sql = ''
    if not is_admin:
        scope_sql = ' AND user_id = :session_user_id'
        params['session_user_id'] = session.get('user_id')

    try:
        with engine.connect() as conn:
            row = conn.execute(text(f"""
                SELECT id, user_id, username, adminpanel_username, observation, filter_json,
                       clients_json, total_success, created_at
                FROM databridge_web.databridge_card_hygiene_logs
                WHERE id = :log_id {scope_sql}
                LIMIT 1
            """), params).mappings().first()

        if not row:
            return jsonify({'error': 'Lote nao encontrado.'}), 404

        created_at = row.get('created_at')
        if created_at is not None and not isinstance(created_at, str):
            created_at = created_at.strftime('%d/%m/%Y %H:%M')

        return jsonify({
            'id': int(row.get('id') or 0),
            'username': row.get('username') or 'Desconhecido',
            'adminpanel_username': row.get('adminpanel_username') or '-',
            'observation': row.get('observation') or '',
            'filter_summary': summarize_hygiene_filter(row.get('filter_json')),
            'total_success': int(row.get('total_success') or 0),
            'created_at': created_at or '',
            'clients': sanitize_hygiene_history_clients(row.get('clients_json')),
            'scope': 'all' if is_admin else 'own',
        }), 200
    except Exception as e:
        print('  [HYGIENE_HISTORY] Erro em /api/hygiene_history/<id>:')
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/hygiene_history/<int:log_id>/export', methods=['GET'])
@login_required
def export_hygiene_history_detail(log_id):
    permission_error = require_hygiene_dashboard_permission()
    if permission_error:
        return permission_error

    current_user = get_current_user()
    is_admin = bool(current_user and current_user.get('is_admin'))
    params = {'log_id': log_id}
    scope_sql = ''
    if not is_admin:
        scope_sql = ' AND user_id = :session_user_id'
        params['session_user_id'] = session.get('user_id')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        with engine.connect() as conn:
            row = conn.execute(text(f"""
                SELECT id, user_id, username, adminpanel_username, observation, filter_json,
                       clients_json, total_success, created_at
                FROM databridge_web.databridge_card_hygiene_logs
                WHERE id = :log_id {scope_sql}
                LIMIT 1
            """), params).mappings().first()

            if not row:
                return jsonify({'error': 'Lote nao encontrado.'}), 404

            job_result = find_hygiene_job_result_for_log(conn, row)

        clients = sanitize_hygiene_history_clients(row.get('clients_json'))
        failed_items = job_result.get('failed_items') or []
        skipped_items = job_result.get('skipped_items') or []
        created_at = row.get('created_at')
        created_label = created_at.strftime('%d/%m/%Y %H:%M') if hasattr(created_at, 'strftime') else str(created_at or '')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sucessos'

        header_fill = PatternFill('solid', fgColor='1F2937')
        header_font = Font(color='FFFFFF', bold=True)
        title_font = Font(size=14, bold=True, color='111827')
        label_font = Font(bold=True, color='374151')
        border = Border(bottom=Side(style='thin', color='D1D5DB'))

        ws['A1'] = f'Lote de Higienização #{row.get("id")}'
        ws['A1'].font = title_font
        ws.merge_cells('A1:F1')

        metadata = [
            ('Data', created_label),
            ('Operador Databridge', row.get('username') or 'Desconhecido'),
            ('Operador AdminPanel', row.get('adminpanel_username') or '-'),
            ('Quantidade de cartões', int(row.get('total_success') or len(clients) or 0)),
            ('Falhas', int(job_result.get('failure_count') or len(failed_items) or 0)),
            ('Ignorados', int(job_result.get('skip_count') or len(skipped_items) or 0)),
            ('Observação', row.get('observation') or ''),
            ('Filtro utilizado', summarize_hygiene_filter(row.get('filter_json'))),
        ]
        for idx, (label, value) in enumerate(metadata, start=3):
            ws.cell(row=idx, column=1, value=label).font = label_font
            ws.cell(row=idx, column=2, value=value)
            ws.cell(row=idx, column=2).alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells('B9:F9')
        ws.merge_cells('B10:F10')

        header_row = 12
        columns = ['Nome', 'CPF', 'Cartão', 'Observação automática', 'Nascimento', 'Fonte nascimento']
        for col_idx, label in enumerate(columns, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row_idx, client in enumerate(clients, start=header_row + 1):
            values = [
                client.get('nome') or '',
                client.get('cpf') or '',
                client.get('cartao') or '',
                client.get('note') or '',
                client.get('birthdate') or '',
                client.get('birthdate_source') or '',
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if col_idx == 2:
                    cell.number_format = '@'

        widths = [34, 16, 22, 58, 16, 22]
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
        ws.freeze_panes = 'A13'

        def add_result_sheet(title, items, extra_columns):
            sheet = wb.create_sheet(title)
            sheet['A1'] = f'{title} do lote #{row.get("id")}'
            sheet['A1'].font = title_font
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4 + len(extra_columns))

            headers = ['Nome', 'CPF', 'Cartão', 'Observação automática'] + [label for label, _ in extra_columns]
            for col_idx, label in enumerate(headers, start=1):
                cell = sheet.cell(row=3, column=col_idx, value=label)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            for row_idx, item in enumerate(items, start=4):
                values = [
                    item.get('nome') or '',
                    item.get('cpf') or '',
                    item.get('cartao') or '',
                    item.get('note') or '',
                ] + [getter(item) for _, getter in extra_columns]
                for col_idx, value in enumerate(values, start=1):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    if col_idx == 2:
                        cell.number_format = '@'

            sheet_widths = [34, 16, 22, 58] + [58 for _ in extra_columns]
            for col_idx, width in enumerate(sheet_widths, start=1):
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
            sheet.freeze_panes = 'A4'

        add_result_sheet('Falhas', failed_items, [
            ('Erro', lambda item: item.get('error') or ''),
            ('Nascimento', lambda item: item.get('birthdate') or ''),
            ('Fonte nascimento', lambda item: item.get('birthdate_source') or ''),
        ])
        add_result_sheet('Ignorados', skipped_items, [
            ('Status', lambda item: item.get('status') or ''),
            ('Motivo', lambda item: item.get('reason') or ''),
            ('Nascimento', lambda item: item.get('birthdate') or ''),
            ('Fonte nascimento', lambda item: item.get('birthdate_source') or ''),
        ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f'historico_higienizacao_lote_{log_id}.xlsx'
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        print('  [HYGIENE_HISTORY] Erro ao exportar lote:')
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/quota_monitoring', methods=['GET'])
@login_required
@permission_required('dashboard')
def get_quota_monitoring():
    """Returns quota-monitoring data on demand based on selected periods and limit."""
    try:
        started_at = time.perf_counter()
        available_periods = [
            {'month': month, 'year': year, 'key': f'{month:02d}-{year}'}
            for month, year in get_quota_available_periods()
        ]
        selected_periods = parse_quota_periods_param(request.args.get('periods', ''), available_periods)
        quota_limit = parse_quota_limit(request.args.get('limit'))
        transaction_profile = parse_quota_transaction_profile(request.args.get('transaction_profile'))
        application_id = parse_quota_application_id(request.args.get('application_id'))
        application_name = QUOTA_APPLICATIONS.get(application_id, 'Escolar')
        summary = get_cached_quota_summary(selected_periods, quota_limit, transaction_profile, application_id)
        cache_hit = summary is not None
        if summary is None:
            summary = fetch_quota_summary_data(selected_periods, quota_limit, transaction_profile, application_id)
            store_cached_quota_summary(selected_periods, quota_limit, summary, transaction_profile, application_id)
        elapsed = time.perf_counter() - started_at
        source_label = 'cache' if cache_hit else 'oracle'
        print(
            f"  [QUOTA] Consulta {source_label} concluida em {elapsed:.2f}s "
            f"(periodos={','.join(selected_periods) or 'nenhum'}, limite={quota_limit:.2f}, perfil={transaction_profile}, app={application_id})"
        )
        return jsonify({
            'last_updated': datetime.datetime.now().strftime('%d/%m/%Y %H:%M'),
            'limit': quota_limit,
            'transaction_profile': transaction_profile,
            'application_id': application_id,
            'application_name': application_name,
            'applications': [
                {'id': app_id, 'label': label}
                for app_id, label in QUOTA_APPLICATIONS.items()
            ],
            'available_periods': available_periods,
            'selected_periods': selected_periods,
            'mode': 'live_query',
            'cache_hit': cache_hit,
            'ranking': summary['ranking'],
            'daily': [],
            'totals': summary['totals']
        }), 200
    except Exception as e:
        print('  [QUOTA] Erro em /api/quota_monitoring:')
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/quota_monitoring/daily', methods=['GET'])
@login_required
@permission_required('dashboard')
def get_quota_monitoring_daily():
    """Returns chart data separately so the main quota panel can load faster."""
    try:
        started_at = time.perf_counter()
        available_periods = [
            {'month': month, 'year': year, 'key': f'{month:02d}-{year}'}
            for month, year in get_quota_available_periods()
        ]
        selected_periods = parse_quota_periods_param(request.args.get('periods', ''), available_periods)
        quota_limit = parse_quota_limit(request.args.get('limit'))
        transaction_profile = parse_quota_transaction_profile(request.args.get('transaction_profile'))
        application_id = parse_quota_application_id(request.args.get('application_id'))
        application_name = QUOTA_APPLICATIONS.get(application_id, 'Escolar')
        daily = get_cached_quota_daily(selected_periods, quota_limit, transaction_profile, application_id)
        cache_hit = daily is not None
        if daily is None:
            daily = fetch_quota_daily_data(selected_periods, quota_limit, transaction_profile, application_id)
            store_cached_quota_daily(selected_periods, quota_limit, daily, transaction_profile, application_id)
        elapsed = time.perf_counter() - started_at
        print(
            f"  [QUOTA] Grafico diario {'cache' if cache_hit else 'oracle'} concluido em {elapsed:.2f}s "
            f"(periodos={','.join(selected_periods) or 'nenhum'}, limite={quota_limit:.2f}, perfil={transaction_profile}, app={application_id})"
        )
        return jsonify({
            'limit': quota_limit,
            'transaction_profile': transaction_profile,
            'application_id': application_id,
            'application_name': application_name,
            'selected_periods': selected_periods,
            'cache_hit': cache_hit,
            'daily': daily
        }), 200
    except Exception as e:
        print('  [QUOTA] Erro em /api/quota_monitoring/daily:')
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/quota_monitoring/user_details', methods=['GET'])
@login_required
@permission_required('dashboard')
def get_quota_monitoring_user_details():
    """Returns detailed transactions only for the selected quota-monitoring user."""
    try:
        available_periods = [
            {'month': month, 'year': year, 'key': f'{month:02d}-{year}'}
            for month, year in get_quota_available_periods()
        ]
        selected_periods = parse_quota_periods_param(request.args.get('periods', ''), available_periods)
        quota_limit = parse_quota_limit(request.args.get('limit'))
        transaction_profile = parse_quota_transaction_profile(request.args.get('transaction_profile'))
        application_id = parse_quota_application_id(request.args.get('application_id'))
        application_name = QUOTA_APPLICATIONS.get(application_id, 'Escolar')
        user_id = request.args.get('user_id', '').strip()
        if not user_id.isdigit():
            return jsonify({'error': 'Usuario invalido.'}), 400

        details = fetch_quota_user_details(int(user_id), selected_periods, quota_limit, application_id, transaction_profile)
        has_more_info = fetch_quota_user_has_more_details(int(user_id), selected_periods, quota_limit, application_id, transaction_profile)
        return jsonify({
            'user_id': int(user_id),
            'limit': quota_limit,
            'transaction_profile': transaction_profile,
            'application_id': application_id,
            'application_name': application_name,
            'selected_periods': selected_periods,
            'has_more_info': has_more_info,
            'data': details
        }), 200
    except Exception as e:
        print('  [QUOTA] Erro em /api/quota_monitoring/user_details:')
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/quota_monitoring/export', methods=['GET'])
@login_required
@permission_required('dashboard')
def export_quota_monitoring():
    """Export quota monitoring with summary and detailed list based on selected periods."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        available_periods = [
            {'month': month, 'year': year, 'key': f'{month:02d}-{year}'}
            for month, year in get_quota_available_periods()
        ]
        selected_periods = parse_quota_periods_param(request.args.get('periods', ''), available_periods)
        quota_limit = parse_quota_limit(request.args.get('limit'))
        transaction_profile = parse_quota_transaction_profile(request.args.get('transaction_profile'))
        application_id = parse_quota_application_id(request.args.get('application_id'))
        application_name = QUOTA_APPLICATIONS.get(application_id, 'Escolar')
        filtered_data = fetch_quota_data(selected_periods, quota_limit, transaction_profile, application_id)
        summary_rows = build_quota_summary_rows(filtered_data, quota_limit)

        total_clientes = len(summary_rows)
        total_transacoes = len(filtered_data)
        total_compras = sum(1 for item in filtered_data if is_quota_purchase_transaction(item))
        total_transferencias = sum(1 for item in filtered_data if is_quota_transfer_transaction(item))
        total_lancamentos_adm = sum(1 for item in filtered_data if item.get('status_movimento_key') == 'R')
        total_valor = round(sum(float(item.get('valor_transacao') or 0) for item in filtered_data), 2)
        total_excedente = round(sum(row['total_excedente'] for row in summary_rows), 2)
        periodos_str = ', '.join(selected_periods) if selected_periods else 'Nenhum periodo'
        atualizado_em = datetime.datetime.now().strftime('%d/%m/%Y %H:%M')

        wb = openpyxl.Workbook()
        ws_resumo = wb.active
        ws_resumo.title = 'Resumo'
        ws_lista = wb.create_sheet('Lista')

        title_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
        section_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
        header_fill = PatternFill(start_color='1D4ED8', end_color='1D4ED8', fill_type='solid')
        alt_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        white_font = Font(color='FFFFFF', bold=True)
        title_font = Font(color='93C5FD', bold=True, size=13)
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        ws_resumo.merge_cells('A1:L1')
        ws_resumo['A1'] = 'Monitoramento de Cotas'
        ws_resumo['A1'].font = title_font
        ws_resumo['A1'].fill = title_fill
        ws_resumo['A1'].alignment = Alignment(horizontal='center')
        ws_resumo['A2'] = f'Periodos selecionados: {periodos_str}'
        ws_resumo['A3'] = f'Limite considerado: R$ {quota_limit:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
        ws_resumo['A4'] = f'Aplicacao selecionada: {application_name} ({application_id})'
        ws_resumo['A5'] = f'Perfil selecionado: {transaction_profile}'
        ws_resumo['A6'] = f'Dados atualizados em: {atualizado_em}'

        metric_rows = [
            ('Clientes excedentes', total_clientes),
            ('Valor total comprado', total_valor),
            ('Total de transacoes', total_transacoes),
            ('Compras classificadas', total_compras),
            ('Transferencias classificadas', total_transferencias),
            ('Lancamentos administrativos', total_lancamentos_adm),
            ('Total excedente', total_excedente),
        ]
        ws_resumo['A7'] = 'Resumo Geral'
        ws_resumo['A7'].font = white_font
        ws_resumo['A7'].fill = section_fill

        for idx, (label, value) in enumerate(metric_rows, start=8):
            ws_resumo[f'A{idx}'] = label
            ws_resumo[f'B{idx}'] = value
            ws_resumo[f'A{idx}'].border = thin_border
            ws_resumo[f'B{idx}'].border = thin_border

        ranking_title_row = 8 + len(metric_rows) + 1
        header_row = ranking_title_row + 1

        ws_resumo[f'A{ranking_title_row}'] = 'Ranking Consolidado'
        ws_resumo[f'A{ranking_title_row}'].font = white_font
        ws_resumo[f'A{ranking_title_row}'].fill = section_fill

        resumo_headers = [
            'Posicao', 'Nome', 'CPF', 'Cartoes', 'Periodos', 'Qtd Periodos',
            'Qtd Transacoes', 'Qtd Compras', 'Qtd Transferencias', 'Qtd Lanc. Adm.', 'Classificacao',
            'Total Comprado', 'Total Excedente'
        ]
        for col_idx, header in enumerate(resumo_headers, 1):
            cell = ws_resumo.cell(row=header_row, column=col_idx, value=header)
            cell.font = white_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for row_idx, row in enumerate(summary_rows, start=header_row + 1):
            values = [
                row_idx - header_row,
                row['nome'],
                row['cpf'],
                row['cartoes'],
                row['periodos'],
                row['qtd_periodos'],
                row['qtd_transacoes'],
                row['qtd_compras'],
                row['qtd_transferencias'],
                row['qtd_lancamentos_adm'],
                row['classificacao_resumo'],
                row['total_comprado'],
                row['total_excedente'],
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws_resumo.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        lista_headers = [
            'Periodo', 'Data Hora', 'USR_ID', 'CPF', 'Nome', 'Cartao',
            'Tipo', 'Status Movimento', 'Cartao Origem Transferencia', 'Valor',
            'Total Acumulado Periodo', 'Excedente Periodo'
        ]
        for col_idx, header in enumerate(lista_headers, 1):
            cell = ws_lista.cell(row=1, column=col_idx, value=header)
            cell.font = white_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        sorted_list = sorted(
            filtered_data,
            key=lambda item: (
                item.get('periodo') or '',
                item.get('usr_name') or '',
                item.get('data_hora_compra') or ''
            )
        )
        for row_idx, item in enumerate(sorted_list, start=2):
            total_periodo = float(item.get('total_acumulado_estouro') or 0)
            values = [
                item.get('periodo', ''),
                item.get('data_hora_compra', ''),
                item.get('usr_id', ''),
                item.get('cpf', ''),
                item.get('usr_name', ''),
                item.get('cartao', ''),
                item.get('tipo_transacao', ''),
                item.get('status_movimento', ''),
                item.get('cartao_origem_transferencia', ''),
                float(item.get('valor_transacao') or 0),
                total_periodo,
                max(total_periodo - quota_limit, 0),
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws_lista.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        for ws in (ws_resumo, ws_lista):
            for col_idx in range(1, ws.max_column + 1):
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(col_idx)
                for row_idx in range(1, ws.max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    cell_value = '' if cell_value is None else str(cell_value)
                    max_length = max(max_length, len(cell_value))
                ws.column_dimensions[column_letter].width = min(max_length + 3, 40)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M')
        filename = f'cotas_monitoramento_{timestamp}.xlsx'
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print('  [QUOTA] Erro em /api/quota_monitoring/export:')
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/attention_users/export', methods=['GET'])
@admin_required
def export_attention_users():
    """Export attention users to XLSX."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        cache = load_cached_json(ATTENTION_CACHE, refresh_attention_cache)
        users = cache.get('users', [])
        last_updated = cache.get('last_updated', '')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Cadastros Incompletos'

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin', color='334155'),
            right=Side(style='thin', color='334155'),
            top=Side(style='thin', color='334155'),
            bottom=Side(style='thin', color='334155')
        )

        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = f'Cadastros Incompletos - Ultimos 7 Dias (Atualizado: {last_updated})'
        title_cell.font = Font(bold=True, size=13, color='F59E0B')
        title_cell.fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        headers = ['Tipo', 'Pendencia', 'Registrado Por', 'Registro Cad', 'Data Criacao']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        alt_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        for row_idx, user in enumerate(users, 4):
            faltando = []
            if not user.get('email'):
                faltando.append('E-mail')
            if not user.get('telefone'):
                faltando.append('Celular')
            faltando_str = ' e '.join(faltando)

            ws.cell(row=row_idx, column=1, value=user.get('tipo', ''))
            ws.cell(row=row_idx, column=2, value=faltando_str)
            ws.cell(row=row_idx, column=3, value=user.get('registrado_por', ''))
            ws.cell(row=row_idx, column=4, value=user.get('registro_cad', ''))
            ws.cell(row=row_idx, column=5, value=user.get('data_criacao', ''))

            if row_idx % 2 == 0:
                for col in range(1, 6):
                    ws.cell(row=row_idx, column=col).fill = alt_fill

        for col_idx in range(1, 6):
            max_len = len(headers[col_idx - 1])
            for row_idx in range(4, len(users) + 4):
                cell_val = str(ws.cell(row=row_idx, column=col_idx).value or '')
                if len(cell_val) > max_len:
                    max_len = len(cell_val)
            ws.column_dimensions[chr(64 + col_idx)].width = min(max_len + 4, 45)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'cadastros_incompletos_{datetime.datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/hygiene_dashboard/export', methods=['GET'])
@login_required
def export_hygiene_dashboard():
    permission_error = require_hygiene_dashboard_permission()
    if permission_error:
        return permission_error

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        payload = build_hygiene_dashboard_payload()
        totals = payload.get('totals') or {}
        ranking = payload.get('ranking') or []
        daily = payload.get('daily') or []
        details = payload.get('details') or []

        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = 'Resumo'
        ws_logs = wb.create_sheet('Lotes')

        title_fill = PatternFill(start_color='062B3D', end_color='062B3D', fill_type='solid')
        section_fill = PatternFill(start_color='0F3B54', end_color='0F3B54', fill_type='solid')
        header_fill = PatternFill(start_color='0EA5E9', end_color='0EA5E9', fill_type='solid')
        alt_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        white_font = Font(color='FFFFFF', bold=True)
        title_font = Font(color='67E8F9', bold=True, size=13)
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        ws_summary.merge_cells('A1:F1')
        ws_summary['A1'] = 'Estatistica de Higienizacao'
        ws_summary['A1'].font = title_font
        ws_summary['A1'].fill = title_fill
        ws_summary['A1'].alignment = Alignment(horizontal='center')
        ws_summary['A2'] = f"Atualizado em: {payload.get('last_updated', '')}"

        metric_rows = [
            ('Total de cartoes higienizados', int(totals.get('total_cartoes') or 0)),
            ('Total de lotes', int(totals.get('total_lotes') or 0)),
            ('Operadores', int(totals.get('total_operadores') or 0)),
            ('Cartoes hoje', int(totals.get('cartoes_hoje') or 0)),
            ('Cartoes ultimos 7 dias', int(totals.get('cartoes_7d') or 0)),
        ]

        ws_summary['A4'] = 'Resumo Geral'
        ws_summary['A4'].font = white_font
        ws_summary['A4'].fill = section_fill

        for idx, (label, value) in enumerate(metric_rows, start=5):
            ws_summary[f'A{idx}'] = label
            ws_summary[f'B{idx}'] = value
            ws_summary[f'A{idx}'].border = thin_border
            ws_summary[f'B{idx}'].border = thin_border

        ranking_title_row = 12
        ws_summary[f'A{ranking_title_row}'] = 'Ranking por Operador'
        ws_summary[f'A{ranking_title_row}'].font = white_font
        ws_summary[f'A{ranking_title_row}'].fill = section_fill

        ranking_headers = ['Posicao', 'Usuario', 'Cartoes', 'Lotes', 'Ultima Execucao']
        for col_idx, header in enumerate(ranking_headers, start=1):
            cell = ws_summary.cell(row=ranking_title_row + 1, column=col_idx, value=header)
            cell.font = white_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for row_idx, item in enumerate(ranking, start=ranking_title_row + 2):
            values = [
                row_idx - (ranking_title_row + 1),
                item.get('username', ''),
                int(item.get('total_cartoes') or 0),
                int(item.get('total_lotes') or 0),
                item.get('ultima_execucao', ''),
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        daily_start_row = ranking_title_row + max(len(ranking), 1) + 4
        ws_summary[f'A{daily_start_row}'] = 'Ocorrencia Diaria'
        ws_summary[f'A{daily_start_row}'].font = white_font
        ws_summary[f'A{daily_start_row}'].fill = section_fill

        daily_headers = ['Data', 'Cartoes', 'Lotes']
        for col_idx, header in enumerate(daily_headers, start=1):
            cell = ws_summary.cell(row=daily_start_row + 1, column=col_idx, value=header)
            cell.font = white_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for row_idx, item in enumerate(daily, start=daily_start_row + 2):
            values = [
                item.get('date', ''),
                int(item.get('total_cartoes') or 0),
                int(item.get('total_lotes') or 0),
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        detail_headers = ['ID Lote', 'Usuario', 'Cartoes', 'Data', 'Observacao']
        for col_idx, header in enumerate(detail_headers, start=1):
            cell = ws_logs.cell(row=1, column=col_idx, value=header)
            cell.font = white_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for row_idx, item in enumerate(details, start=2):
            values = [
                int(item.get('id') or 0),
                item.get('username', ''),
                int(item.get('total_cartoes') or 0),
                item.get('created_at', ''),
                item.get('observation', ''),
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws_logs.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        for ws in (ws_summary, ws_logs):
            for col_idx in range(1, ws.max_column + 1):
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(col_idx)
                for row_idx in range(1, ws.max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    cell_value = '' if cell_value is None else str(cell_value)
                    max_length = max(max_length, len(cell_value))
                ws.column_dimensions[column_letter].width = min(max_length + 3, 45)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M')
        filename = f'higienizacao_dashboard_{timestamp}.xlsx'
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print('  [HYGIENE_DASHBOARD] Erro em /api/hygiene_dashboard/export:')
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def _build_estudantes_xlsx(rows, title_text, tab_name, accent_color):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tab_name

    header_font  = Font(bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill(start_color=accent_color, end_color=accent_color, fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border  = Border(
        left=Side(style='thin', color='334155'), right=Side(style='thin', color='334155'),
        top=Side(style='thin', color='334155'),  bottom=Side(style='thin', color='334155')
    )
    alt_fill = PatternFill(start_color='F5F3FF', end_color='F5F3FF', fill_type='solid')

    ws.merge_cells('A1:E1')
    tc = ws['A1']
    tc.value = f'{title_text} — gerado em {datetime.datetime.now().strftime("%d/%m/%Y %H:%M")}'
    tc.font  = Font(bold=True, size=12, color=accent_color)
    tc.fill  = PatternFill(start_color='F5F3FF', end_color='F5F3FF', fill_type='solid')
    tc.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    headers = ['CPF', 'Nome', 'Última Requisição', 'Data Início', 'Data Término']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = header_align; cell.border = thin_border

    for row_idx, row in enumerate(rows, 4):
        vals = [row['cpf'], row['nome'], row['ultima_requisicao'], row['data_inicio'], row['data_termino']]
        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val or '')
            cell.border = thin_border; cell.alignment = Alignment(vertical='center')
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    for col_idx, width in enumerate([18, 40, 20, 16, 16], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return out


@dashboard_bp.route('/api/dashboard/anomalias_estudantis', methods=['GET'])
@login_required
def anomalias_estudantis():
    """Retorna contagens para os dois cards de anomalia estudantil no dashboard."""
    try:
        with engine.connect() as conn:
            sem_legacydb_count = conn.execute(text("""
                SELECT COUNT(DISTINCT REGEXP_REPLACE(va.cpf, '[^0-9]', ''))
                FROM databridge_db.vw_alunos_aprovados va
                LEFT JOIN sntr_interligar.SALES_CAD_UNICO_JSON scuj
                    ON REGEXP_REPLACE(va.cpf, '[^0-9]', '') =
                       REPLACE(REPLACE(CONVERT(scuj.cpf USING utf8mb4) COLLATE utf8mb4_unicode_ci, '.', ''), '-', '')
                WHERE va.status = 'Aprovado'
                  AND scuj.cpf IS NULL
            """)).scalar() or 0

            sem_cartao_count = conn.execute(text("""
                SELECT COUNT(DISTINCT REGEXP_REPLACE(va.cpf, '[^0-9]', ''))
                FROM databridge_db.vw_alunos_aprovados va
                INNER JOIN sntr_interligar.SALES_CAD_UNICO_JSON scuj
                    ON REGEXP_REPLACE(va.cpf, '[^0-9]', '') =
                       REPLACE(REPLACE(CONVERT(scuj.cpf USING utf8mb4) COLLATE utf8mb4_unicode_ci, '.', ''), '-', '')
                WHERE va.status = 'Aprovado'
                  AND (
                      scuj.cartoes_json IS NULL
                      OR scuj.cartoes_json = 'null'
                      OR scuj.cartoes_json = '{}'
                      OR scuj.cartoes_json NOT LIKE '%\"58.03.%'
                  )
            """)).scalar() or 0

        return jsonify({'sem_legacydb': int(sem_legacydb_count), 'sem_cartao': int(sem_cartao_count)}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/dashboard/estudantes_sem_cartao/export', methods=['GET'])
@login_required
def export_estudantes_sem_cartao():
    """Exporta aprovados que existem no LegacyDB mas não têm cartão tipo 58.03."""
    try:
        with engine.connect() as conn:
            rows = conn.execute(text("""
                SELECT
                    REGEXP_REPLACE(va.cpf, '[^0-9]', '') AS cpf,
                    va.nome,
                    DATE_FORMAT(MAX(va.data_requisicao), '%d/%m/%Y') AS ultima_requisicao,
                    DATE_FORMAT(MAX(va.data_inicio),     '%d/%m/%Y') AS data_inicio,
                    DATE_FORMAT(MAX(va.data_termino),    '%d/%m/%Y') AS data_termino
                FROM databridge_db.vw_alunos_aprovados va
                INNER JOIN sntr_interligar.SALES_CAD_UNICO_JSON scuj
                    ON REGEXP_REPLACE(va.cpf, '[^0-9]', '') =
                       REPLACE(REPLACE(CONVERT(scuj.cpf USING utf8mb4) COLLATE utf8mb4_unicode_ci, '.', ''), '-', '')
                WHERE va.status = 'Aprovado'
                  AND (
                      scuj.cartoes_json IS NULL
                      OR scuj.cartoes_json = 'null'
                      OR scuj.cartoes_json = '{}'
                      OR scuj.cartoes_json NOT LIKE '%\"58.03.%'
                  )
                GROUP BY REGEXP_REPLACE(va.cpf, '[^0-9]', ''), va.nome
                ORDER BY MAX(va.data_requisicao) DESC
            """)).mappings().fetchall()

        out = _build_estudantes_xlsx(
            rows,
            title_text='Aprovados no LegacyDB sem Cartão Estudantil (58.03)',
            tab_name='Sem Cartão 58.03',
            accent_color='0EA5E9'
        )
        filename = f'aprovados_sem_cartao_{datetime.datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@dashboard_bp.route('/api/dashboard/estudantes_sem_legacydb/export', methods=['GET'])
@login_required
def export_estudantes_sem_legacydb():
    """Exporta aprovados que não existem no LegacyDB."""
    try:
        with engine.connect() as conn:
            rows = conn.execute(text("""
                SELECT
                    REGEXP_REPLACE(va.cpf, '[^0-9]', '') AS cpf,
                    va.nome,
                    DATE_FORMAT(MAX(va.data_requisicao), '%d/%m/%Y') AS ultima_requisicao,
                    DATE_FORMAT(MAX(va.data_inicio),     '%d/%m/%Y') AS data_inicio,
                    DATE_FORMAT(MAX(va.data_termino),    '%d/%m/%Y') AS data_termino
                FROM databridge_db.vw_alunos_aprovados va
                LEFT JOIN sntr_interligar.SALES_CAD_UNICO_JSON scuj
                    ON REGEXP_REPLACE(va.cpf, '[^0-9]', '') =
                       REPLACE(REPLACE(CONVERT(scuj.cpf USING utf8mb4) COLLATE utf8mb4_unicode_ci, '.', ''), '-', '')
                WHERE va.status = 'Aprovado'
                  AND scuj.cpf IS NULL
                GROUP BY REGEXP_REPLACE(va.cpf, '[^0-9]', ''), va.nome
                ORDER BY MAX(va.data_requisicao) DESC
            """)).mappings().fetchall()

        out = _build_estudantes_xlsx(
            rows,
            title_text='Aprovados no Portal Estudante sem cadastro no LegacyDB',
            tab_name='Sem LegacyDB',
            accent_color='7C3AED'
        )
        filename = f'aprovados_sem_legacydb_{datetime.datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
