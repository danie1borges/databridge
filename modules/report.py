import os
import re
import time
import uuid
import json
import threading
import io
import datetime
import pandas as pd
from flask import Blueprint, request, jsonify, send_file, session
from werkzeug.utils import secure_filename, safe_join
from sqlalchemy import text
from core.config import UPLOAD_FOLDER
from core.database import engine
from modules.auth import login_required, permission_required, user_has_permission
from modules.card_hygiene import (
    ensure_card_hygiene_tables,
    get_card_hygiene_exclusion_sql,
    persist_card_hygiene_log,
    cancel_card_hygiene_task,
    get_card_hygiene_progress,
    CARD_HYGIENE_MAX_BATCH,
    CARD_HYGIENE_DAILY_QUOTA,
    get_user_daily_hygiene_usage,
    create_hygiene_job,
    start_hygiene_background_job,
    get_job_from_db,
    get_active_jobs_for_user,
    get_all_active_jobs,
    request_cancel_job,
    set_hygiene_job_popup_closed,
    validate_vtadmin_credentials,
)
from modules.contact_fallbacks import delivery_fallback_query

report_bp = Blueprint('report', __name__)

REPORT_FILTER_FIELDS = {
    'cpf': {'column': 'f.cpf', 'type': 'string'},
    'cartao': {'column': 'f.cartao', 'type': 'string'},
    'tipo_cartao': {'column': 'f.tipo_cartao', 'type': 'select'},
    'app_id': {'column': 'f.app_id', 'type': 'select'},
    'saldo': {'column': 'f.saldo', 'type': 'number'},
    'valor_ultima_recarga': {'column': 'f.valor_ultima_recarga', 'type': 'number'},
    'valor_recarga_pendente': {'column': 'f.valor_recarga_pendente', 'type': 'number'},
    'ultimo_uso': {'column': 'f.ultimo_uso', 'type': 'date'},
    'ultima_recarga': {'column': 'f.ultima_recarga', 'type': 'date'},
    'recarga_pendente': {'column': 'f.recarga_pendente', 'type': 'date'},
    'ultima_compra_data': {'column': None, 'type': 'derived_netsales'},
    'aluno_sem_direito': {'column': None, 'type': 'flag'},
    'aprovado_sem_cartao_estudantil': {'column': None, 'type': 'flag'},
}

REPORT_FILTER_OPERATORS = {
    'string': {'contains', 'equals', 'starts_with', 'ends_with', 'is_empty', 'is_not_empty'},
    'number': {'equals', 'gt', 'gte', 'lt', 'lte', 'between', 'is_empty', 'is_not_empty'},
    'date': {'on', 'after', 'before', 'between', 'is_empty', 'is_not_empty'},
    'derived_netsales': {'on', 'after', 'before', 'between', 'is_empty', 'is_not_empty'},
    'select': {'equals', 'not_equals', 'is_empty', 'is_not_empty'},
    'flag': {'is_true'}
}

REPORT_APP_NAMES = {
    '400': 'VALE TRANSPORTE',
    '500': 'COMUM',
    '600': 'DEFICIENTE FISICO',
    '610': 'DEFICIENTE C/AC',
    '611': 'ACOMPANHANTE',
    '900': 'ESCOLAR GRATUIDADE',
    '700': 'FUNCIONAL',
    '800': 'IDOSO',
    '301': 'FISCAL TRANSPORTE',
    '302': 'FUNC CONS TUTELAR',
    '303': 'FUNC POLICIA CIVIL',
    '304': 'FUNC EMTU',
    '305': 'FUNC DATACROSS',
    '306': 'FUNC TEMPORARIO 03',
    '308': 'FUNC TEMP 02',
    '309': 'INSS - EMPRESAS',
    '310': 'FUNC SISTEMA',
    '605': 'ESP CADEIRANTE',
    '910': 'ESCOLAR',
    '620': 'ESPECIAL (NAO CATRACA)',
    '625': 'ESPECIAL C/AC (NAO CATRACA)',
    '312': 'FUNCIONARIO IBGE',
    '505': 'P SOCIAL',
    '260': 'MANUTENCAO',
    '905': 'ESCOLAR GRATUIDADE MUNICIPAL'
}

# Scalar subquery: last COM_NETSALES purchase date for f.cartao (e.g. "58.03.06348712-1").
# Parses ISS_ID / CD_ID / CRD_SNR from the card string using SUBSTRING_INDEX.
_NETSALES_LAST_SALE_SQL = """(
    SELECT MAX(n.DATA)
    FROM sntr_interligar.COM_NETSALES n
    WHERE n.ISS_ID  = CAST(SUBSTRING_INDEX(f.cartao, '.', 1) AS UNSIGNED)
      AND n.CD_ID   = CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(f.cartao, '.', 2), '.', -1) AS UNSIGNED)
      AND n.CRD_SNR = CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(f.cartao, '-', 1), '.', -1) AS UNSIGNED)
)"""

REPORT_FILTERS_CACHE_TTL = 600
_report_filters_cache = {
    'tipos_cartao': None,
    'apps': None,
    'updated_at': 0
}
_report_filters_cache_lock = threading.Lock()

_REPORT_LIVE_SOURCE_SQL_TEMPLATE = """
    SELECT
        j.cpf,
        cards.card_key AS cartao,
        tipos.tipo_key AS tipo_cartao,
        apps.app_key AS app_id,
        CAST(JSON_UNQUOTE(JSON_EXTRACT(j.cartoes_json, CONCAT('$.\"', cards.card_key, '\".\"', tipos.tipo_key, '\".\"', apps.app_key, '\".saldo'))) AS DECIMAL(10,2)) AS saldo,
        STR_TO_DATE(JSON_UNQUOTE(JSON_EXTRACT(j.cartoes_json, CONCAT('$.\"', cards.card_key, '\".\"', tipos.tipo_key, '\".\"', apps.app_key, '\".ultimo_uso'))), '%Y-%m-%d %H:%i:%s') AS ultimo_uso,
        STR_TO_DATE(JSON_UNQUOTE(JSON_EXTRACT(j.cartoes_json, CONCAT('$.\"', cards.card_key, '\".\"', tipos.tipo_key, '\".\"', apps.app_key, '\".ultima_recarga'))), '%Y-%m-%d %H:%i:%s') AS ultima_recarga,
        CAST(JSON_UNQUOTE(JSON_EXTRACT(j.cartoes_json, CONCAT('$.\"', cards.card_key, '\".\"', tipos.tipo_key, '\".\"', apps.app_key, '\".valor_ultima_recarga'))) AS DECIMAL(10,2)) AS valor_ultima_recarga,
        STR_TO_DATE(JSON_UNQUOTE(JSON_EXTRACT(j.cartoes_json, CONCAT('$.\"', cards.card_key, '\".\"', tipos.tipo_key, '\".\"', apps.app_key, '\".recarga_pendente'))), '%Y-%m-%d %H:%i:%s') AS recarga_pendente,
        CAST(JSON_UNQUOTE(JSON_EXTRACT(j.cartoes_json, CONCAT('$.\"', cards.card_key, '\".\"', tipos.tipo_key, '\".\"', apps.app_key, '\".valor_recarga_pendente'))) AS DECIMAL(10,2)) AS valor_recarga_pendente
    FROM sntr_interligar.SALES_CAD_UNICO_JSON j
    JOIN JSON_TABLE(JSON_KEYS(j.cartoes_json), '$[*]' COLUMNS (card_key VARCHAR(100) PATH '$')) cards
    JOIN JSON_TABLE(JSON_KEYS(JSON_EXTRACT(j.cartoes_json, CONCAT('$.\"', cards.card_key, '\"'))), '$[*]' COLUMNS (tipo_key VARCHAR(100) PATH '$')) tipos
    JOIN JSON_TABLE(JSON_KEYS(JSON_EXTRACT(j.cartoes_json, CONCAT('$.\"', cards.card_key, '\".\"', tipos.tipo_key, '\"'))), '$[*]' COLUMNS (app_key VARCHAR(100) PATH '$')) apps
    {app_key_filter}
"""
# Keep the no-filter variant as default for backwards compat.
REPORT_LIVE_SOURCE_SQL = _REPORT_LIVE_SOURCE_SQL_TEMPLATE.format(app_key_filter='')


def _extract_app_id_filter(filters):
    """Returns list of numeric app_id strings from direct AND-level app_id equals rules.
    Returns None when no safe push-down can be determined."""
    if not filters or not isinstance(filters, dict):
        return None
    if str(filters.get('condition', 'AND')).upper() != 'AND':
        return None
    app_ids = []
    for rule in (filters.get('rules') or []):
        if not isinstance(rule, dict) or 'rules' in rule:
            continue
        if rule.get('field') == 'app_id' and rule.get('operator') == 'equals':
            val = str(rule.get('value') or '').strip()
            if re.match(r'^\d+$', val):
                app_ids.append(val)
    return app_ids if app_ids else None


def _get_live_source_sql(app_ids=None):
    """Returns live_flat source SQL, optionally filtered by app_key in the CTE."""
    if not app_ids:
        return REPORT_LIVE_SOURCE_SQL
    ids_sql = ', '.join(f"'{aid}'" for aid in app_ids)
    return _REPORT_LIVE_SOURCE_SQL_TEMPLATE.format(
        app_key_filter=f"WHERE apps.app_key IN ({ids_sql})"
    )

def ensure_report_preferences_table(conn):
    conn.execute(text("CREATE DATABASE IF NOT EXISTS datacross_web DEFAULT CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci"))
    conn.execute(text("""
        CREATE TABLE IF NOT EXISTS datacross_web.datacross_user_report_filters (
            user_id INT NOT NULL,
            report_key VARCHAR(100) NOT NULL,
            filters_json LONGTEXT NULL,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
            PRIMARY KEY (user_id, report_key),
            CONSTRAINT fk_datacross_user_report_filters_user
                FOREIGN KEY (user_id) REFERENCES datacross_web.datacross_users(id) ON DELETE CASCADE
        )
    """))


def get_saved_report_builder_state(conn, user_id, report_key='clientes_cartoes_ativos'):
    ensure_report_preferences_table(conn)
    row = conn.execute(text("""
        SELECT filters_json
        FROM datacross_web.datacross_user_report_filters
        WHERE user_id = :user_id AND report_key = :report_key
    """), {'user_id': user_id, 'report_key': report_key}).fetchone()

    if not row or not row[0]:
        return None

    try:
        return json.loads(row[0])
    except Exception:
        return None


def save_report_builder_state(conn, user_id, builder_state, report_key='clientes_cartoes_ativos'):
    ensure_report_preferences_table(conn)
    conn.execute(text("""
        INSERT INTO datacross_web.datacross_user_report_filters (user_id, report_key, filters_json)
        VALUES (:user_id, :report_key, :filters_json)
        ON DUPLICATE KEY UPDATE filters_json = VALUES(filters_json), updated_at = CURRENT_TIMESTAMP
    """), {
        'user_id': user_id,
        'report_key': report_key,
        'filters_json': json.dumps(builder_state, ensure_ascii=False)
    })


def get_cached_report_filter_options(conn):
    now = time.time()
    with _report_filters_cache_lock:
        if (
            _report_filters_cache['apps'] is not None
            and _report_filters_cache['tipos_cartao'] is not None
            and now - _report_filters_cache['updated_at'] < REPORT_FILTERS_CACHE_TTL
        ):
            return _report_filters_cache['apps'], _report_filters_cache['tipos_cartao']

    tipos_cartao = conn.execute(text(f"""
        WITH live_flat AS (
            {REPORT_LIVE_SOURCE_SQL}
        )
        SELECT DISTINCT tipo_cartao
        FROM live_flat
        WHERE tipo_cartao IS NOT NULL AND tipo_cartao != ''
        ORDER BY tipo_cartao
    """)).fetchall()

    apps = [
        {'id': app_id, 'label': label}
        for app_id, label in sorted(REPORT_APP_NAMES.items(), key=lambda item: int(item[0]))
    ]
    tipos = [r[0] for r in tipos_cartao]

    with _report_filters_cache_lock:
        _report_filters_cache['apps'] = apps
        _report_filters_cache['tipos_cartao'] = tipos
        _report_filters_cache['updated_at'] = now

    return apps, tipos


def warm_report_filters_cache():
    try:
        with engine.connect() as conn:
            get_cached_report_filter_options(conn)
        print('  [REPORT] Cache de filtros do relatorio aquecido.')
    except Exception as e:
        print(f'  [REPORT] Falha ao aquecer cache de filtros: {e}')


def user_can_manage_card_hygiene():
    return user_has_permission('higienizacao')


def report_filters_include_field(node, field_id):
    if not node or not isinstance(node, dict):
        return False
    for rule in node.get('rules') or []:
        if not isinstance(rule, dict):
            continue
        if rule.get('field') == field_id:
            return True
        if report_filters_include_field(rule, field_id):
            return True
    return False


def prepare_alunos_sem_direito_temp_table(conn, filters):
    needs_temp = (
        report_filters_include_field(filters, 'aluno_sem_direito') or
        report_filters_include_field(filters, 'aprovado_sem_cartao_estudantil')
    )
    if not needs_temp:
        return False
    if conn.info.get('alunos_sem_direito_temp_ready'):
        return True

    conn.execute(text("""
        CREATE TEMPORARY TABLE IF NOT EXISTS tmp_datacross_alunos_aprovados (
            cpf VARCHAR(20) NOT NULL,
            modalidade VARCHAR(16) NOT NULL,
            PRIMARY KEY (cpf, modalidade)
        )
    """))
    conn.execute(text("TRUNCATE TABLE tmp_datacross_alunos_aprovados"))
    conn.execute(text("""
        INSERT IGNORE INTO tmp_datacross_alunos_aprovados (cpf, modalidade)
        SELECT DISTINCT
            REPLACE(REPLACE(cpf, '.', ''), '-', '') AS cpf,
            CASE
                WHEN LOWER(TRIM(modalidade)) LIKE 'gratuit%' THEN 'GRATUITO'
                ELSE 'MEIA'
            END AS modalidade
        FROM datacross_db.vw_alunos_aprovados
        WHERE cpf IS NOT NULL AND cpf != ''
    """))
    conn.info['alunos_sem_direito_temp_ready'] = True
    return True


def build_report_where(filters, params):
    where_clauses = ["1=1", get_card_hygiene_exclusion_sql('f')]
    filter_sql = build_report_filter_clause(filters, params, [1])
    if filter_sql:
        where_clauses.append(f"({filter_sql})")
    return " AND ".join(where_clauses)


def build_report_where_without_hygiene_exclusion(filters, params):
    where_clauses = ["1=1"]
    filter_sql = build_report_filter_clause(filters, params, [1])
    if filter_sql:
        where_clauses.append(f"({filter_sql})")
    return " AND ".join(where_clauses)


def release_reactivated_report_cards(conn, filters):
    ensure_card_hygiene_tables(conn)
    prepare_alunos_sem_direito_temp_table(conn, filters)
    params = {}
    where_sql = build_report_where_without_hygiene_exclusion(filters, params)
    live_source_sql = _get_live_source_sql(_extract_app_id_filter(filters))
    rows = conn.execute(text(f"""
        WITH live_flat AS (
            {live_source_sql}
        )
        SELECT DISTINCT cartao
        FROM live_flat f
        WHERE {where_sql}
    """), params).fetchall()

    card_numbers = [str(row[0]).strip() for row in rows if row and str(row[0] or '').strip()]
    if not card_numbers:
        return 0

    updated = 0
    chunk_size = 500
    for start in range(0, len(card_numbers), chunk_size):
        chunk = card_numbers[start:start + chunk_size]
        update_params = {}
        placeholders = []
        for idx, card_number in enumerate(chunk):
            key = f"react_card_{start}_{idx}"
            placeholders.append(f":{key}")
            update_params[key] = card_number

        result = conn.execute(text(f"""
            UPDATE datacross_web.datacross_card_hygiene_hidden_cards
            SET is_active = 0,
                reactivated_at = COALESCE(reactivated_at, CURRENT_TIMESTAMP),
                updated_at = CURRENT_TIMESTAMP
            WHERE is_active = 1
              AND CONVERT(card_number USING utf8mb4) COLLATE utf8mb4_0900_ai_ci IN ({', '.join(placeholders)})
        """), update_params)
        updated += int(result.rowcount or 0)

    return updated


def try_release_reactivated_report_cards(conn, filters):
    try:
        # Keep this cleanup out of the report transaction. If a hygiene row is
        # locked by another process, the report must not wait for MySQL's
        # default lock timeout.
        with engine.connect() as release_conn:
            release_conn.execute(text("SET SESSION innodb_lock_wait_timeout = 1"))
            updated = release_reactivated_report_cards(release_conn, filters)
            release_conn.commit()
            return updated
    except Exception as exc:
        print(f"[RELATORIO] Ignorando reativacao de cartoes por lock/erro temporario: {exc}")
        return 0


_SORT_FIELD_MAP = {
    'cpf':              'f.cpf',
    'cartao':           'f.cartao',
    'uso':              'f.ultimo_uso',
    'recarga':          'f.ultima_recarga',
    'recarga_pendente': 'f.recarga_pendente',
    'saldo':            'f.saldo',
}

def _inject_netsales_cte(where_sql, params):
    """If where_sql references nsl_last_sale, inject the netsales CTE into the query.
    Returns (extra_cte_sql, live_table_name) — caller substitutes into the WITH block.
    nsl_last_sale = COALESCE(COM_NETSALES.last_sale, Mercury.ultima_recarga):
    COM_NETSALES is always preferred; Mercury is used only when COM_NETSALES has no record."""
    params.pop('__nsl_min__', None)
    if 'nsl_last_sale' not in where_sql:
        return '', 'live_flat'
    cte = """
        , nsl_agg AS (
            SELECT ISS_ID, CD_ID, CRD_SNR, MAX(DATA) AS last_sale
            FROM sntr_interligar.COM_NETSALES
            GROUP BY ISS_ID, CD_ID, CRD_SNR
        )
        , live_nsl AS (
            SELECT f.*,
                   COALESCE(na.last_sale, f.ultima_recarga) AS nsl_last_sale
            FROM live_flat f
            LEFT JOIN nsl_agg na ON
                na.ISS_ID  = CAST(SUBSTRING_INDEX(f.cartao, '.', 1) AS UNSIGNED)
                AND na.CD_ID   = CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(f.cartao, '.', 2), '.', -1) AS UNSIGNED)
                AND na.CRD_SNR = CAST(SUBSTRING_INDEX(SUBSTRING_INDEX(f.cartao, '-', 1), '.', -1) AS UNSIGNED)
        )"""
    return cte, 'live_nsl'


def fetch_report_rows(conn, filters, limit=None, offset=None, sort_by='cpf', sort_dir='asc'):
    prepare_alunos_sem_direito_temp_table(conn, filters)
    params = {}
    where_sql = build_report_where(filters, params)
    netsales_cte, live_table = _inject_netsales_cte(where_sql, params)

    live_source_sql = _get_live_source_sql(_extract_app_id_filter(filters))
    sort_col = _SORT_FIELD_MAP.get(sort_by, 'f.cpf')
    order_dir = 'DESC' if str(sort_dir).lower() == 'desc' else 'ASC'
    order_clause = f"{sort_col} {order_dir}"
    if sort_col != 'f.cpf':
        order_clause += ', f.cpf ASC'

    base_query = f"""
        WITH live_flat AS (
            {live_source_sql}
        ){netsales_cte}
        SELECT f.cpf, f.cartao, f.tipo_cartao, f.app_id, f.saldo, f.ultimo_uso, f.ultima_recarga,
               f.valor_ultima_recarga, f.recarga_pendente, f.valor_recarga_pendente,
               CASE
                   WHEN hc.card_number IS NOT NULL
                        AND hc.is_active = 1
                   THEN 1 ELSE 0
               END AS higienizacao_historico,
               CASE
                   WHEN hc.reactivated_at IS NOT NULL
                        AND COALESCE(hc.is_active, 0) = 0
                        AND hc.reactivated_at >= DATE_SUB(NOW(), INTERVAL 1 DAY)
                   THEN 1 ELSE 0
               END AS higienizacao_recente,
               hc.reactivated_at AS higienizacao_reativada_em
        FROM {live_table} f
        LEFT JOIN datacross_web.datacross_card_hygiene_hidden_cards hc
            ON CONVERT(hc.card_number USING utf8mb4) COLLATE utf8mb4_0900_ai_ci =
               CONVERT(f.cartao USING utf8mb4) COLLATE utf8mb4_0900_ai_ci
        WHERE {where_sql}
        ORDER BY {order_clause}
    """
    if limit is not None:
        base_query += "\nLIMIT :limit"
        params['limit'] = int(limit)
        if offset is not None:
            base_query += " OFFSET :offset"
            params['offset'] = int(offset)
    return conn.execute(text(base_query), params).mappings().fetchall()


def augment_results_with_contacts(flat_rows, conn):
    data_list = []
    if not flat_rows:
        return data_list

    raw_cpfs = [r['cpf'] for r in flat_rows]
    formatted_cpfs = [f"{c[:3]}.{c[3:6]}.{c[6:9]}-{c[9:]}" for c in raw_cpfs if len(c) == 11]
    all_formats = list(set(raw_cpfs + formatted_cpfs))

    cpf_placeholders = ', '.join([f":cpf_{i}" for i in range(len(all_formats))])
    cpf_params = {f"cpf_{i}": all_formats[i] for i in range(len(all_formats))}

    delivery_sql = delivery_fallback_query(f"u.cpf IN ({cpf_placeholders})").text
    cliente_q = f"""
        SELECT
            REPLACE(REPLACE(c.cpf, '.', ''), '-', '') as num_cpf,
            c.name,
            c.email,
            COALESCE(dl.celular_entrega, c.cellphone) as cellphone,
            COALESCE(dl.endereco_entrega, '') as endereco
        FROM sntr_cliente.customer c
        LEFT JOIN (
            {delivery_sql}
        ) dl ON dl.cpf_limpo = REPLACE(REPLACE(c.cpf, '.', ''), '-', '')
        WHERE c.cpf IN ({cpf_placeholders})
    """
    clientes = {r['num_cpf']: dict(r) for r in conn.execute(text(cliente_q), cpf_params).mappings().fetchall()}

    mercury_q = f"SELECT REPLACE(REPLACE(cpf, '.', ''), '-', '') as num_cpf, nome, email, telefone as cellphone FROM sntr_interligar.SALES_CAD_UNICO_JSON WHERE cpf IN ({cpf_placeholders})"
    mercurys = {r['num_cpf']: dict(r) for r in conn.execute(text(mercury_q), cpf_params).mappings().fetchall()}

    abt_q = f"""
        SELECT num_cpf, nome, email, celular as cellphone
        FROM (
            SELECT REPLACE(REPLACE(documento, '.', ''), '-', '') as num_cpf, nome, email, celular,
                   ROW_NUMBER() OVER (PARTITION BY REPLACE(REPLACE(documento, '.', ''), '-', '') ORDER BY data_cadastro DESC) as rn
            FROM sntr_interligar.COM_CLIENTES_ABT
            WHERE documento IN ({cpf_placeholders})
        ) sub WHERE rn = 1
    """
    abts = {r['num_cpf']: dict(r) for r in conn.execute(text(abt_q), cpf_params).mappings().fetchall()}

    estudante_q = f"SELECT cpf as num_cpf, email, celular as cellphone FROM datacross_db.alunos WHERE cpf IN ({cpf_placeholders}) GROUP BY cpf"
    estudantes = {r['num_cpf']: dict(r) for r in conn.execute(text(estudante_q), cpf_params).mappings().fetchall()}

    wifi_q = f"SELECT CPF as num_cpf, MAX(EMAIL) as email, MAX(TELEFONE) as cellphone FROM sntr_interligar.WIFIMAX_USERS WHERE CPF IN ({cpf_placeholders}) GROUP BY CPF"
    wifis = {r['num_cpf']: dict(r) for r in conn.execute(text(wifi_q), cpf_params).mappings().fetchall()}

    whatsapp_q = f"SELECT DISTINCT cpf as num_cpf, telefone as cellphone FROM sntr_interligar.CLIENTES_WHATSAPP WHERE cpf IN ({cpf_placeholders})"
    whatsapps = {r['num_cpf']: dict(r) for r in conn.execute(text(whatsapp_q), cpf_params).mappings().fetchall()}

    for r in flat_rows:
        d = dict(r)
        cpf = d['cpf']

        nome = None
        if cpf in mercurys and mercurys[cpf].get('nome'):
            nome = mercurys[cpf]['nome']
        elif cpf in clientes and clientes[cpf].get('name'):
            nome = clientes[cpf]['name']
        elif cpf in abts and abts[cpf].get('nome'):
            nome = abts[cpf]['nome']

        email, celular, origem = None, None, 'MERCURY_ONLY'
        if cpf in mercurys and (mercurys[cpf].get('email') or mercurys[cpf].get('cellphone')):
            email, celular, origem = mercurys[cpf].get('email'), mercurys[cpf].get('cellphone'), 'MERCURY_APP'
        elif cpf in clientes and (clientes[cpf].get('email') or clientes[cpf].get('cellphone')):
            email, celular, origem = clientes[cpf].get('email'), clientes[cpf].get('cellphone'), 'CLIENTE'
        elif cpf in abts and (abts[cpf].get('email') or abts[cpf].get('cellphone')):
            email, celular, origem = abts[cpf].get('email'), abts[cpf].get('cellphone'), 'ABT'
        elif cpf in estudantes and (estudantes[cpf].get('email') or estudantes[cpf].get('cellphone')):
            email, celular, origem = estudantes[cpf].get('email'), estudantes[cpf].get('cellphone'), 'ESTUDANTE'
        elif cpf in wifis and (wifis[cpf].get('email') or wifis[cpf].get('cellphone')):
            email, celular, origem = wifis[cpf].get('email'), wifis[cpf].get('cellphone'), 'WIFI'
        elif cpf in whatsapps and whatsapps[cpf].get('cellphone'):
            celular, origem = whatsapps[cpf].get('cellphone'), 'WHATSAPP'

        d['nome'] = nome
        d['email'] = email
        d['celular'] = celular
        d['origem_contato'] = origem
        d['endereco_cliente'] = clientes.get(cpf, {}).get('endereco', '')

        for k, v in d.items():
            if v is not None and not isinstance(v, (int, str, float, bool)):
                d[k] = str(v)
        data_list.append(d)

    return data_list


def enrich_with_last_sale(data_list, conn):
    """Adiciona ultima_compra_data/local/valor de COM_NETSALES em cada row do relatório."""
    import datetime as _dt
    from modules.search import parse_card_number_parts, normalize_sale_origin

    key_to_rows = {}
    for row in data_list:
        card = str(row.get('cartao') or '').strip()
        parts = parse_card_number_parts(card)
        if parts:
            key = (parts['iss_id'], parts['cd_id'], parts['crd_snr'])
            key_to_rows.setdefault(key, []).append(row)

    if not key_to_rows:
        return

    params = {}
    conditions = []
    for i, (iss_id, cd_id, crd_snr) in enumerate(key_to_rows):
        conditions.append(f'(n.ISS_ID = :iss{i} AND n.CD_ID = :cd{i} AND n.CRD_SNR = :snr{i})')
        params[f'iss{i}'] = iss_id
        params[f'cd{i}'] = cd_id
        params[f'snr{i}'] = crd_snr

    sql = text(f"""
        SELECT ISS_ID, CD_ID, CRD_SNR, PONTO_VENDA, DATA, VALOR
        FROM (
            SELECT ISS_ID, CD_ID, CRD_SNR, PONTO_VENDA, DATA, VALOR,
                   ROW_NUMBER() OVER (PARTITION BY ISS_ID, CD_ID, CRD_SNR ORDER BY DATA DESC) AS rn
            FROM sntr_interligar.COM_NETSALES n
            WHERE {' OR '.join(conditions)}
        ) ranked
        WHERE rn = 1
    """)

    try:
        results = conn.execute(sql, params).mappings().fetchall()
    except Exception as exc:
        print(f"[REPORT] Falha ao enriquecer com ultima compra COM_NETSALES: {exc}")
        return

    sale_by_key = {}
    for r in results:
        sale_by_key[(int(r['ISS_ID']), int(r['CD_ID']), int(r['CRD_SNR']))] = r

    for key, rows in key_to_rows.items():
        sale = sale_by_key.get(key)
        if not sale:
            continue
        raw_dt = sale.get('DATA')
        if isinstance(raw_dt, _dt.datetime):
            data_str = raw_dt.strftime('%d/%m/%Y %H:%M')
        else:
            data_str = str(raw_dt or '')
        try:
            valor_reais = round(float(str(sale.get('VALOR') or '0').replace(',', '.')) / 100, 2)
        except Exception:
            valor_reais = None
        local_str = normalize_sale_origin(sale.get('PONTO_VENDA'))
        for row in rows:
            row['ultima_compra_data'] = data_str
            row['ultima_compra_local'] = local_str
            row['ultima_compra_valor'] = valor_reais

    # Mercury local: preenche apenas rows sem COM_NETSALES (ultima_recarga como fallback).
    missing_netsales = [row for row in data_list if not row.get('ultima_compra_data')]
    if missing_netsales:
        _enrich_report_from_mercury_local(missing_netsales)

    # Reclassifica entradas Mercury que na verdade são transferências de crédito (CAC_KEY1=99999).
    _cross_check_mercury_transfers_report(data_list)

    # Oracle fallback: rows ainda sem ultima_compra_data (transferência de crédito como último recurso).
    missing_oracle = [row for row in data_list if not row.get('ultima_compra_data')]
    if missing_oracle:
        _enrich_report_from_cardaccount(missing_oracle)


def _enrich_report_from_mercury_local(rows):
    """Compara ultima_recarga do Mercury (SALES_CAD_UNICO_JSON) com a data já encontrada
    e sobrescreve se Mercury for mais recente. Usa CPF como chave (mais confiável que JSON_CONTAINS_PATH)."""
    import datetime as _dt
    from modules.search import parse_local_card_datetime
    from core.database import engine, safe_json_parse
    from sqlalchemy import text as _text

    if not rows:
        return

    raw_cpfs = list({str(row.get('cpf') or '').strip() for row in rows if row.get('cpf')})
    if not raw_cpfs:
        return

    formatted_cpfs = [f"{c[:3]}.{c[3:6]}.{c[6:9]}-{c[9:]}" for c in raw_cpfs if len(c) == 11]
    all_cpfs = list(set(raw_cpfs + formatted_cpfs))
    placeholders = ', '.join(f':c{i}' for i in range(len(all_cpfs)))
    params = {f'c{i}': cpf for i, cpf in enumerate(all_cpfs)}

    try:
        with engine.connect() as conn:
            db_rows = conn.execute(_text(f"""
                SELECT REPLACE(REPLACE(cpf, '.', ''), '-', '') AS cpf_norm, cartoes_json
                FROM sntr_interligar.SALES_CAD_UNICO_JSON
                WHERE cpf IN ({placeholders})
            """), params).mappings().fetchall()
    except Exception as exc:
        print(f"[REPORT] Falha ao consultar Mercury local: {exc}")
        return

    mercury_by_cpf = {}
    for db_row in db_rows:
        cpf_norm = str(db_row.get('cpf_norm') or '').strip()
        parsed = safe_json_parse(db_row['cartoes_json'])
        if isinstance(parsed, dict):
            mercury_by_cpf[cpf_norm] = parsed

    for row in rows:
        cpf_norm = str(row.get('cpf') or '').strip().replace('.', '').replace('-', '')
        card = str(row.get('cartao') or '').strip()
        row_app_id = str(row.get('app_id') or '').strip()
        parsed = mercury_by_cpf.get(cpf_norm)
        if not isinstance(parsed, dict):
            continue
        card_data = parsed.get(card)
        if not isinstance(card_data, dict):
            continue

        best_dt = None
        best_value = None
        for card_type, apps in card_data.items():
            if not isinstance(apps, dict):
                continue
            for app_id, details in apps.items():
                if row_app_id and app_id != row_app_id:
                    continue
                if not isinstance(details, dict):
                    continue
                recharge_dt = parse_local_card_datetime(details.get('ultima_recarga'))
                if not recharge_dt:
                    continue
                value = details.get('valor_ultima_recarga')
                if best_dt is None or recharge_dt > best_dt:
                    best_dt = recharge_dt
                    best_value = value

        if not best_dt:
            continue

        existing_data = row.get('ultima_compra_data')
        if existing_data:
            try:
                existing_dt = _dt.datetime.strptime(existing_data, '%d/%m/%Y %H:%M')
                if existing_dt >= best_dt.replace(second=0, microsecond=0):
                    continue
            except Exception:
                pass

        row['ultima_compra_data'] = best_dt.strftime('%d/%m/%Y %H:%M')
        row['ultima_compra_local'] = 'Mercury'
        row['ultima_compra_valor'] = best_value


def _cross_check_mercury_transfers_report(rows):
    """Reclassifica para 'Transferência de Crédito' rows onde Mercury foi a fonte mas
    CARDACCOUNT confirma que a recarga veio de uma transferência entre cartões (CAC_KEY1=99999)."""
    import datetime as _dt
    from modules.search import parse_card_number_parts, parse_local_card_datetime

    mercury_rows = [row for row in rows if row.get('ultima_compra_local') == 'Mercury']
    if not mercury_rows:
        return

    check_list = []
    for row in mercury_rows:
        card = str(row.get('cartao') or '').strip()
        parts = parse_card_number_parts(card)
        if not parts:
            continue
        data_str = row.get('ultima_compra_data')
        if not data_str:
            continue
        try:
            mdt = _dt.datetime.strptime(data_str, '%d/%m/%Y %H:%M')
        except Exception:
            continue
        try:
            val_str = str(row.get('ultima_compra_valor') or '0').replace(',', '.')
            vcents = int(round(float(val_str) * 100))
        except Exception:
            continue
        if vcents <= 0:
            continue
        app_id_val = int(row.get('app_id') or 0)
        check_list.append((row, parts['iss_id'], parts['cd_id'], parts['crd_snr'], app_id_val, mdt, vcents))

    if not check_list:
        return

    try:
        from modules.dashboard import get_quota_connection
        oracle = get_quota_connection()
    except Exception as exc:
        print(f"[REPORT] Oracle indisponivel para cruzar transferências Mercury: {exc}")
        return

    try:
        cursor = oracle.cursor()
        for row, iss, cd, snr, app_id, mdt, vcents in check_list:
            d1 = mdt - _dt.timedelta(days=3)
            d2 = mdt + _dt.timedelta(days=3)
            cursor.execute("""
                SELECT COUNT(*) FROM CARDACCOUNT c
                WHERE c.ISS_ID = :iss AND c.CD_ID = :cd AND c.CRD_SNR = :snr
                  AND c.APP_ID = :app_id
                  AND c.CAC_TYPE = 'C' AND c.CAC_STATUS = 'C'
                  AND c.CAC_TRANVALUE = :val
                  AND COALESCE(c.CAC_EFECTDATE, c.CAC_TRANDATE) BETWEEN :d1 AND :d2
                  AND (
                      c.CAC_KEY1 = 99999
                      OR EXISTS (
                          SELECT 1 FROM CARDACCOUNT d
                          WHERE d.USR_ID = c.USR_ID
                            AND d.APP_ID = c.APP_ID
                            AND d.CAC_TRANDATE = c.CAC_TRANDATE
                            AND d.CAC_TRANVALUE = c.CAC_TRANVALUE
                            AND d.CAC_TYPE = 'D'
                            AND (d.CI_ID <> c.CI_ID OR d.CRD_SNR <> c.CRD_SNR)
                      )
                  )
            """, {'iss': iss, 'cd': cd, 'snr': snr, 'app_id': app_id, 'val': vcents, 'd1': d1, 'd2': d2})
            if cursor.fetchone()[0]:
                row['ultima_compra_local'] = 'Transferência de Crédito'
    except Exception as exc:
        print(f"[REPORT] Falha ao cruzar transferências Mercury: {exc}")
    finally:
        try:
            oracle.close()
        except Exception:
            pass


def _enrich_report_from_cardaccount(rows):
    """Preenche ultima_compra_data/local/valor via Oracle CARDACCOUNT para rows sem COM_NETSALES."""
    import datetime as _dt
    from modules.search import parse_card_number_parts
    try:
        from modules.dashboard import get_quota_connection
        oracle = get_quota_connection()
    except Exception as exc:
        print(f"[REPORT] Oracle indisponivel para fallback CARDACCOUNT: {exc}")
        return

    key_to_rows = {}
    for row in rows:
        card = str(row.get('cartao') or '').strip()
        parts = parse_card_number_parts(card)
        if parts:
            app_id_val = int(row.get('app_id') or 0)
            key = (parts['iss_id'], parts['cd_id'], parts['crd_snr'], app_id_val)
            key_to_rows.setdefault(key, []).append(row)

    if not key_to_rows:
        try:
            oracle.close()
        except Exception:
            pass
        return

    try:
        conditions = ' OR '.join(
            f'(c.ISS_ID = :iss{i} AND c.CD_ID = :cd{i} AND c.CRD_SNR = :snr{i} AND c.APP_ID = :app{i})'
            for i in range(len(key_to_rows))
        )
        params = {}
        for i, (iss, cd, snr, app_id) in enumerate(key_to_rows):
            params[f'iss{i}'] = iss
            params[f'cd{i}']  = cd
            params[f'snr{i}'] = snr
            params[f'app{i}'] = app_id
        cursor = oracle.cursor()
        cursor.execute(f"""
            SELECT ISS_ID, CD_ID, CRD_SNR, APP_ID, CAC_TRANDATE, CAC_EFECTDATE, CAC_TRANVALUE, CAC_KEY1, CAC_STATUS
            FROM (
                SELECT c.ISS_ID, c.CD_ID, c.CRD_SNR, c.APP_ID, c.CAC_TRANDATE, c.CAC_EFECTDATE,
                       c.CAC_TRANVALUE, c.CAC_KEY1, c.CAC_STATUS,
                       ROW_NUMBER() OVER (PARTITION BY c.ISS_ID, c.CD_ID, c.CRD_SNR, c.APP_ID
                                         ORDER BY COALESCE(c.CAC_EFECTDATE, c.CAC_TRANDATE) DESC,
                                                  c.CAC_SEQNBR DESC) AS rn
                FROM CARDACCOUNT c
                WHERE c.CAC_TYPE = 'C'
                  AND c.CAC_STATUS = 'C'
                  AND ({conditions})
                  AND (
                      c.CAC_KEY1 = 99999
                      OR EXISTS (
                          SELECT 1 FROM CARDACCOUNT d
                          WHERE d.USR_ID = c.USR_ID
                            AND d.APP_ID = c.APP_ID
                            AND d.CAC_TRANDATE = c.CAC_TRANDATE
                            AND d.CAC_TRANVALUE = c.CAC_TRANVALUE
                            AND d.CAC_TYPE = 'D'
                            AND (d.CI_ID <> c.CI_ID OR d.CRD_SNR <> c.CRD_SNR)
                      )
                  )
            )
            WHERE rn = 1
        """, params)
        cols = [d[0] for d in cursor.description]
        results = [dict(zip(cols, r)) for r in cursor.fetchall()]
    except Exception as exc:
        print(f"[REPORT] Falha ao consultar CARDACCOUNT em lote: {exc}")
        return
    finally:
        try:
            oracle.close()
        except Exception:
            pass

    for r in results:
        key = (int(r['ISS_ID']), int(r['CD_ID']), int(r['CRD_SNR']), int(r['APP_ID']))
        target_rows = key_to_rows.get(key, [])
        raw_dt = r.get('CAC_EFECTDATE') or r.get('CAC_TRANDATE')
        data_str = raw_dt.strftime('%d/%m/%Y %H:%M') if isinstance(raw_dt, _dt.datetime) else str(raw_dt or '')
        local_str = 'Transferência de Crédito'
        try:
            valor_reais = round(float(r.get('CAC_TRANVALUE') or 0) / 100, 2)
        except Exception:
            valor_reais = None
        for row in target_rows:
            row['ultima_compra_data'] = data_str
            row['ultima_compra_local'] = local_str
            row['ultima_compra_valor'] = valor_reais


def build_hygiene_preview_items(flat_rows, conn):
    if not flat_rows:
        return []

    cpfs = []
    seen_cpfs = set()
    for row in flat_rows:
        cpf = str(row.get('cpf') or '').strip()
        if cpf and cpf not in seen_cpfs:
            seen_cpfs.add(cpf)
            cpfs.append(cpf)

    names_by_cpf = {}
    if cpfs:
        placeholders = ', '.join(f":cpf_{i}" for i in range(len(cpfs)))
        params = {f"cpf_{i}": cpfs[i] for i in range(len(cpfs))}
        for raw_name, raw_cpf in conn.execute(text(f"""
            SELECT nome, REPLACE(REPLACE(cpf, '.', ''), '-', '') AS num_cpf
            FROM sntr_interligar.SALES_CAD_UNICO_JSON
            WHERE REPLACE(REPLACE(cpf, '.', ''), '-', '') IN ({placeholders})
        """), params).fetchall():
            cpf_key = str(raw_cpf or '').strip()
            if cpf_key and cpf_key not in names_by_cpf:
                names_by_cpf[cpf_key] = raw_name or ''

    return [{
        'cpf': row.get('cpf'),
        'nome': names_by_cpf.get(str(row.get('cpf') or '').strip(), ''),
        'cartao': row.get('cartao'),
        'app_id': row.get('app_id'),
        'tipo_cartao': row.get('tipo_cartao'),
        'higienizacao_historico': bool(row.get('higienizacao_historico')),
        'higienizacao_recente': bool(row.get('higienizacao_recente')),
    } for row in flat_rows]


def build_report_filter_clause(node, params, counter):
    if not node or not isinstance(node, dict):
        return None

    rules = node.get('rules') or []
    condition = 'OR' if str(node.get('condition', 'AND')).upper() == 'OR' else 'AND'
    sql_parts = []

    for rule in rules:
        if not isinstance(rule, dict):
            continue

        if 'rules' in rule:
            nested_clause = build_report_filter_clause(rule, params, counter)
            if nested_clause:
                sql_parts.append(f"({nested_clause})")
            continue

        field_id = rule.get('field')
        field_cfg = REPORT_FILTER_FIELDS.get(field_id)
        if not field_cfg:
            continue

        operator = rule.get('operator')
        if operator not in REPORT_FILTER_OPERATORS[field_cfg['type']]:
            continue

        # derived_netsales: references the nsl_last_sale column injected by the netsales CTE.
        # The CTE is injected by fetch_report_rows / count query when this sentinel is present.
        if field_cfg['type'] == 'derived_netsales':
            col = 'f.nsl_last_sale'
            value = rule.get('value') or ''
            value_to = rule.get('value_to') or ''
            if operator == 'is_empty':
                sql_parts.append(f"{col} IS NULL")
            elif operator == 'is_not_empty':
                sql_parts.append(f"{col} IS NOT NULL")
            elif operator in ('on', 'after', 'before', 'between'):
                if value in (None, ''):
                    continue
                if operator == 'on':
                    p1 = f"rf_{counter[0]}"; counter[0] += 1
                    p2 = f"rf_{counter[0]}"; counter[0] += 1
                    params[p1] = f"{value} 00:00:00"
                    params[p2] = f"{value} 23:59:59"
                    sql_parts.append(f"({col} >= :{p1} AND {col} <= :{p2})")
                elif operator == 'after':
                    p = f"rf_{counter[0]}"; counter[0] += 1
                    params[p] = f"{value} 00:00:00"
                    sql_parts.append(f"{col} >= :{p}")
                elif operator == 'before':
                    p = f"rf_{counter[0]}"; counter[0] += 1
                    params[p] = f"{value} 23:59:59"
                    sql_parts.append(f"{col} <= :{p}")
                elif operator == 'between':
                    if value_to in (None, ''):
                        continue
                    p1 = f"rf_{counter[0]}"; counter[0] += 1
                    p2 = f"rf_{counter[0]}"; counter[0] += 1
                    params[p1] = f"{value} 00:00:00"
                    params[p2] = f"{value_to} 23:59:59"
                    sql_parts.append(f"({col} >= :{p1} AND {col} <= :{p2})")
            continue

        if field_id == 'aluno_sem_direito':
            sql_parts.append("""
                (
                    f.app_id IN ('900', '910')
                    AND NOT EXISTS (
                        SELECT 1
                        FROM tmp_datacross_alunos_aprovados vaa
                        WHERE CONVERT(vaa.cpf USING utf8mb4) COLLATE utf8mb4_0900_ai_ci =
                              CONVERT(REPLACE(REPLACE(f.cpf, '.', ''), '-', '') USING utf8mb4) COLLATE utf8mb4_0900_ai_ci
                          AND (
                              (f.app_id = '910' AND vaa.modalidade IN ('MEIA', 'GRATUITO'))
                              OR
                              (f.app_id = '900' AND vaa.modalidade = 'GRATUITO')
                          )
                    )
                )
            """)
            continue

        if field_id == 'aprovado_sem_cartao_estudantil':
            sql_parts.append("""
                (
                    EXISTS (
                        SELECT 1
                        FROM tmp_datacross_alunos_aprovados vaa
                        WHERE CONVERT(vaa.cpf USING utf8mb4) COLLATE utf8mb4_0900_ai_ci =
                              CONVERT(REPLACE(REPLACE(f.cpf, '.', ''), '-', '') USING utf8mb4) COLLATE utf8mb4_0900_ai_ci
                    )
                    AND NOT EXISTS (
                        SELECT 1
                        FROM sntr_interligar.SALES_CAD_UNICO_JSON j2
                        WHERE REPLACE(REPLACE(CONVERT(j2.cpf USING utf8mb4) COLLATE utf8mb4_0900_ai_ci, '.', ''), '-', '') =
                              REPLACE(REPLACE(f.cpf, '.', ''), '-', '')
                          AND j2.cartoes_json LIKE '%\"58.03.%'
                    )
                )
            """)
            continue

        column = field_cfg['column']

        if operator == 'is_empty':
            if field_cfg['type'] in ('string', 'select'):
                sql_parts.append(f"({column} IS NULL OR {column} = '')")
            else:
                sql_parts.append(f"{column} IS NULL")
            continue
        if operator == 'is_not_empty':
            if field_cfg['type'] in ('string', 'select'):
                sql_parts.append(f"({column} IS NOT NULL AND {column} != '')")
            else:
                sql_parts.append(f"{column} IS NOT NULL")
            continue

        value = rule.get('value')
        value_to = rule.get('value_to')

        if value in (None, ''):
            continue

        if field_cfg['type'] == 'number':
            try:
                value = float(value)
                if operator == 'between':
                    value_to = float(value_to)
            except (TypeError, ValueError):
                continue

        if field_cfg['type'] == 'date':
            if operator == 'between' and value_to in (None, ''):
                continue
            if operator == 'on':
                param_name = f"rf_{counter[0]}"
                counter[0] += 1
                params[param_name] = f"{value} 00:00:00"
                param_name_end = f"rf_{counter[0]}"
                counter[0] += 1
                params[param_name_end] = f"{value} 23:59:59"
                sql_parts.append(f"({column} >= :{param_name} AND {column} <= :{param_name_end})")
                continue
            if operator == 'after':
                param_name = f"rf_{counter[0]}"
                counter[0] += 1
                params[param_name] = f"{value} 00:00:00"
                sql_parts.append(f"{column} >= :{param_name}")
                continue
            if operator == 'before':
                param_name = f"rf_{counter[0]}"
                counter[0] += 1
                params[param_name] = f"{value} 23:59:59"
                sql_parts.append(f"{column} <= :{param_name}")
                continue
            if operator == 'between':
                param_name = f"rf_{counter[0]}"
                counter[0] += 1
                params[param_name] = f"{value} 00:00:00"
                param_name_end = f"rf_{counter[0]}"
                counter[0] += 1
                params[param_name_end] = f"{value_to} 23:59:59"
                sql_parts.append(f"({column} >= :{param_name} AND {column} <= :{param_name_end})")
                continue

        param_name = f"rf_{counter[0]}"
        counter[0] += 1

        if operator == 'contains':
            params[param_name] = f"%{value}%"
            sql_parts.append(f"{column} LIKE :{param_name}")
        elif operator == 'starts_with':
            params[param_name] = f"{value}%"
            sql_parts.append(f"{column} LIKE :{param_name}")
        elif operator == 'ends_with':
            params[param_name] = f"%{value}"
            sql_parts.append(f"{column} LIKE :{param_name}")
        elif operator == 'equals':
            params[param_name] = value
            sql_parts.append(f"{column} = :{param_name}")
        elif operator == 'not_equals':
            params[param_name] = value
            sql_parts.append(f"{column} != :{param_name}")
        elif operator == 'gt':
            params[param_name] = value
            sql_parts.append(f"{column} > :{param_name}")
        elif operator == 'gte':
            params[param_name] = value
            sql_parts.append(f"{column} >= :{param_name}")
        elif operator == 'lt':
            params[param_name] = value
            sql_parts.append(f"{column} < :{param_name}")
        elif operator == 'lte':
            params[param_name] = value
            sql_parts.append(f"{column} <= :{param_name}")
        elif operator == 'between':
            if value_to in (None, ''):
                continue
            params[param_name] = value
            param_name_end = f"rf_{counter[0]}"
            counter[0] += 1
            params[param_name_end] = value_to
            sql_parts.append(f"({column} >= :{param_name} AND {column} <= :{param_name_end})")

    return f" {condition} ".join(sql_parts) if sql_parts else None


def _has_post_date_filters(node):
    """Returns True if the filter tree contains any post_date rules."""
    if not node or not isinstance(node, dict):
        return False
    for rule in (node.get('rules') or []):
        if not isinstance(rule, dict):
            continue
        if 'rules' in rule:
            if _has_post_date_filters(rule):
                return True
        else:
            cfg = REPORT_FILTER_FIELDS.get(rule.get('field') or '')
            if cfg and cfg['type'] == 'post_date':
                return True
    return False


def _collect_post_date_rules(node):
    """Flatten all post_date rules from the filter tree into a list of dicts."""
    rules = []
    if not node or not isinstance(node, dict):
        return rules
    for rule in (node.get('rules') or []):
        if not isinstance(rule, dict):
            continue
        if 'rules' in rule:
            rules.extend(_collect_post_date_rules(rule))
        else:
            cfg = REPORT_FILTER_FIELDS.get(rule.get('field') or '')
            if cfg and cfg['type'] == 'post_date':
                rules.append(rule)
    return rules


def apply_post_date_filters(data_list, filters):
    """Filter data_list in Python based on post_date rules (e.g. ultima_compra_data).
    Returns a new list containing only rows that satisfy ALL post_date rules."""
    import datetime as _dt

    rules = _collect_post_date_rules(filters)
    if not rules:
        return data_list

    def parse_row_dt(row, field):
        raw = row.get(field) or ''
        for fmt in ('%d/%m/%Y %H:%M', '%d/%m/%Y %H:%M:%S', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
            try:
                return _dt.datetime.strptime(str(raw), fmt)
            except Exception:
                pass
        return None

    filtered = []
    for row in data_list:
        ok = True
        for rule in rules:
            field_id = rule.get('field')
            operator = rule.get('operator')
            value = rule.get('value') or ''
            value_to = rule.get('value_to') or ''

            if operator == 'is_empty':
                if parse_row_dt(row, field_id) is not None:
                    ok = False
                continue
            if operator == 'is_not_empty':
                if parse_row_dt(row, field_id) is None:
                    ok = False
                continue

            row_dt = parse_row_dt(row, field_id)

            if operator == 'on':
                try:
                    d = _dt.datetime.strptime(value, '%Y-%m-%d').date()
                    if not (row_dt and row_dt.date() == d):
                        ok = False
                except Exception:
                    ok = False
            elif operator == 'after':
                try:
                    d = _dt.datetime.strptime(value, '%Y-%m-%d')
                    if not (row_dt and row_dt >= d):
                        ok = False
                except Exception:
                    ok = False
            elif operator == 'before':
                try:
                    d = _dt.datetime.strptime(value, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
                    if not (row_dt and row_dt <= d):
                        ok = False
                except Exception:
                    ok = False
            elif operator == 'between':
                try:
                    d1 = _dt.datetime.strptime(value, '%Y-%m-%d')
                    d2 = _dt.datetime.strptime(value_to, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
                    if not (row_dt and d1 <= row_dt <= d2):
                        ok = False
                except Exception:
                    ok = False
        if ok:
            filtered.append(row)
    return filtered


def cleanup_uploads_folder(max_days=3):
    folder = UPLOAD_FOLDER
    if not os.path.exists(folder):
        return

    now = time.time()
    for f in os.listdir(folder):
        p = os.path.join(folder, f)
        if os.path.isfile(p):
            if os.path.getmtime(p) < now - (max_days * 86400):
                try:
                    os.remove(p)
                except Exception as e:
                    print(f"Erro ao deletar arquivo antigo {p}: {e}")


@report_bp.route('/api/relatorio_filters', methods=['GET'])
@login_required
@permission_required('relatorio')
def get_relatorio_filters():
    try:
        with engine.connect() as conn:
            saved_filters = get_saved_report_builder_state(conn, session['user_id'])
            apps, tipos_cartao = get_cached_report_filter_options(conn)
            return jsonify({
                'apps': apps,
                'tipos_cartao': tipos_cartao,
                'saved_filters': saved_filters,
                'can_hygienize': user_can_manage_card_hygiene(),
            }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@report_bp.route('/api/relatorio_filters_state', methods=['POST'])
@login_required
@permission_required('relatorio')
def save_relatorio_filters_state():
    data = request.json or {}
    builder_state = data.get('builder_state')

    try:
        with engine.connect() as conn:
            save_report_builder_state(conn, session['user_id'], builder_state)
            conn.commit()
        return jsonify({'ok': True}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@report_bp.route('/api/relatorio', methods=['POST'])
@login_required
@permission_required('relatorio')
def get_relatorio():
    data = request.json or {}

    page = int(data.get('page', 1))
    per_page = int(data.get('per_page', 15))
    offset = (page - 1) * per_page
    export_excel = data.get('export_excel', False)
    filters = data.get('filters')
    sort_by = data.get('sort_by', 'cpf')
    sort_dir = data.get('sort_dir', 'asc')

    try:
        with engine.connect() as conn:
            conn.info.pop('alunos_sem_direito_temp_ready', None)
            ensure_card_hygiene_tables(conn)
            prepare_alunos_sem_direito_temp_table(conn, filters)
            try_release_reactivated_report_cards(conn, filters)
            params = {}
            where_sql = build_report_where(filters, params)
            count_netsales_cte, count_live_table = _inject_netsales_cte(where_sql, params)
            count_live_source_sql = _get_live_source_sql(_extract_app_id_filter(filters))

            total_records = 0
            if not export_excel:
                count_q = text(f"""
                    WITH live_flat AS (
                        {count_live_source_sql}
                    ){count_netsales_cte}
                    SELECT COUNT(1)
                    FROM {count_live_table} f
                    WHERE {where_sql}
                """)
                total_records = conn.execute(count_q, params).scalar()

            if export_excel:
                flat_rows = fetch_report_rows(conn, filters, limit=20000, sort_by=sort_by, sort_dir=sort_dir)
            else:
                flat_rows = fetch_report_rows(conn, filters, limit=per_page, offset=offset, sort_by=sort_by, sort_dir=sort_dir)
            data_list = augment_results_with_contacts(flat_rows, conn)
            enrich_with_last_sale(data_list, conn)

            # Nome sort is resolved after contact augmentation (page-level)
            if sort_by == 'nome':
                data_list.sort(
                    key=lambda r: (r.get('nome') or '').lower(),
                    reverse=(sort_dir == 'desc')
                )

            if export_excel:
                df = pd.DataFrame(data_list)
                origem_map = {
                    'CLIENTE': 'Área do Cliente',
                    'MERCURY_APP': 'VTWeb Admin',
                    'ABT': 'Cadê meu ônibus recarga',
                    'ESTUDANTE': 'Sou Estudante',
                    'WIFI': 'Wifi Max',
                    'WHATSAPP': 'WhatsApp',
                    'MERCURY_ONLY': 'Apenas Base Física'
                }
                if 'origem_contato' in df.columns:
                    df['origem_contato'] = df['origem_contato'].map(origem_map).fillna('Desconhecido')

                col_order = [
                    'nome', 'cpf', 'email', 'celular', 'endereco_cliente', 'origem_contato',
                    'cartao', 'tipo_cartao', 'app_id', 'saldo', 'ultimo_uso', 'ultima_recarga',
                    'recarga_pendente', 'ultima_compra_data', 'ultima_compra_local', 'ultima_compra_valor'
                ]
                col_headers = {
                    'nome': 'Nome', 'cpf': 'CPF', 'email': 'E-mail', 'celular': 'Celular',
                    'endereco_cliente': 'Endereço', 'origem_contato': 'Fonte do Contato',
                    'cartao': 'Número do Cartão', 'tipo_cartao': 'Tipo do Cartão',
                    'app_id': 'Aplicação (ID)', 'saldo': 'Saldo Disponível',
                    'ultimo_uso': 'Último Uso', 'ultima_recarga': 'Última Recarga',
                    'recarga_pendente': 'Data Recarga Pendente',
                    'ultima_compra_data': 'Última Compra (Data)',
                    'ultima_compra_local': 'Local da Compra',
                    'ultima_compra_valor': 'Valor da Compra (R$)',
                }
                available_cols = [c for c in col_order if c in df.columns]
                df = df[available_cols]
                df.columns = [col_headers[c] for c in available_cols]

                filename = f"Relatorio_{time.strftime('%Y-%m-%d')}.xlsx"
                filepath = os.path.join(UPLOAD_FOLDER, filename)
                df.to_excel(filepath, index=False)

                return jsonify({'download_url': filename}), 200

            return jsonify({
                'data': data_list,
                'total': total_records,
                'page': page,
                'per_page': per_page,
                'total_pages': (total_records + per_page - 1) // per_page
            })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@report_bp.route('/api/relatorio_higienizacao/preview', methods=['POST'])
@login_required
def preview_relatorio_higienizacao():
    if not user_can_manage_card_hygiene():
        return jsonify({'error': 'Você não tem permissão para executar higienização de cadastro.'}), 403

    data = request.json or {}
    filters = data.get('filters')
    if not filters or not (filters.get('rules') or []):
        return jsonify({'error': 'Defina um filtro antes de abrir a higienização de cadastro.'}), 400

    try:
        with engine.connect() as conn:
            conn.info.pop('alunos_sem_direito_temp_ready', None)
            ensure_card_hygiene_tables(conn)
            prepare_alunos_sem_direito_temp_table(conn, filters)
            try_release_reactivated_report_cards(conn, filters)

            count_params = {}
            where_sql = build_report_where(filters, count_params)
            hygiene_live_source_sql = _get_live_source_sql(_extract_app_id_filter(filters))
            total = conn.execute(text(f"""
                WITH live_flat AS (
                    {hygiene_live_source_sql}
                )
                SELECT COUNT(1)
                FROM live_flat f
                WHERE {where_sql}
            """), count_params).scalar() or 0

            # Dynamic quota check
            used_quota = get_user_daily_hygiene_usage(conn, session['user_id'])
            remaining_quota = max(0, CARD_HYGIENE_DAILY_QUOTA - used_quota)

            if total == 0:
                return jsonify({'total': 0, 'items': [], 'max_batch': remaining_quota}), 200

            if remaining_quota <= 0:
                return jsonify({
                    'error': f'Você atingiu sua cota diária de {CARD_HYGIENE_DAILY_QUOTA} cartões. Retorne amanhã.',
                    'total': int(total),
                    'max_batch': 0,
                }), 400

            if total > remaining_quota:
                return jsonify({
                    'error': f'Limite excedido. O relatório tem {total} cartões disponíveis, mas sua cota restante de hoje é de {remaining_quota}.',
                    'total': int(total),
                    'max_batch': remaining_quota,
                }), 400

            flat_rows = fetch_report_rows(conn, filters, limit=remaining_quota)
            items = build_hygiene_preview_items(flat_rows, conn)

            return jsonify({
                'total': int(total),
                'items': items,
                'max_batch': remaining_quota,
            }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@report_bp.route('/api/relatorio_higienizacao/process', methods=['POST'])
@login_required
def process_relatorio_higienizacao():
    if not user_can_manage_card_hygiene():
        return jsonify({'error': 'Você não tem permissão para executar higienização de cadastro.'}), 403

    data = request.json or {}
    filters = data.get('filters')
    selected_cards = [str(card).strip() for card in (data.get('selected_cards') or []) if str(card).strip()]
    observation = (data.get('observation') or '').strip()
    vtadmin_username = str(data.get('vtadmin_username') or '').strip()
    vtadmin_password = data.get('vtadmin_password') or ''

    if not filters or not (filters.get('rules') or []):
        return jsonify({'error': 'Defina um filtro antes de confirmar a higienização.'}), 400
    if not selected_cards:
        return jsonify({'error': 'Selecione pelo menos um cartão para processar.'}), 400
    if not observation:
        return jsonify({'error': 'Informe a observação que será enviada ao VTAdmin.'}), 400

    if not vtadmin_username:
        return jsonify({'error': 'Informe o usuário do VTAdmin para continuar.'}), 400
    if not str(vtadmin_password).strip():
        return jsonify({'error': 'Informe a senha do VTAdmin para continuar.'}), 400

    if len(selected_cards) > CARD_HYGIENE_MAX_BATCH:
        return jsonify({'error': f'Por segurança técnica, processe no máximo {CARD_HYGIENE_MAX_BATCH} cartões por requisição.'}), 400

    user_id = session['user_id']
    username = session.get('username', 'Desconhecido')

    try:
        with engine.connect() as conn:
            conn.info.pop('alunos_sem_direito_temp_ready', None)
            used_quota = get_user_daily_hygiene_usage(conn, user_id)
            remaining_quota = max(0, CARD_HYGIENE_DAILY_QUOTA - used_quota)
            if len(selected_cards) > remaining_quota:
                return jsonify({'error': f'A sua cota diária restante ({remaining_quota}) é insuficiente para higienizar {len(selected_cards)} cartões. Tente na virada do dia.'}), 400

            ensure_card_hygiene_tables(conn)
            try_release_reactivated_report_cards(conn, filters)
            
            # Limiting the maximum selected amount physically possible to process in a batch limit
            candidate_rows = fetch_report_rows(conn, filters, limit=CARD_HYGIENE_MAX_BATCH)
            candidate_data = augment_results_with_contacts(candidate_rows, conn)
            candidate_map = {
                str(item.get('cartao') or '').strip(): {
                    'cpf': str(item.get('cpf') or '').strip(),
                    'nome': item.get('nome') or '',
                    'cartao': str(item.get('cartao') or '').strip(),
                }
                for item in candidate_data
                if str(item.get('cartao') or '').strip()
            }

            selected_items = [candidate_map[card] for card in selected_cards if card in candidate_map]
            missing_cards = [card for card in selected_cards if card not in candidate_map]
            if missing_cards:
                return jsonify({
                    'error': 'Alguns cartões selecionados não continuam elegíveis no filtro atual. Atualize a prévia antes de iniciar.',
                    'cards': missing_cards[:20],
                }), 400
            if not selected_items:
                return jsonify({'error': 'Nenhum cartão selecionado continua elegível no relatório atual.'}), 400

        # Create job record in DB, fire background thread, return immediately
        job_id = create_hygiene_job(user_id, username, observation, selected_cards, len(selected_items), filters)
        start_hygiene_background_job(
            job_id,
            selected_items,
            observation,
            user_id,
            username,
            vtadmin_username,
            vtadmin_password,
            filters,
        )

        return jsonify({
            'ok': True,
            'job_id': job_id,
            'total': len(selected_items),
            'message': f'Higienização de {len(selected_items)} cartões iniciada em log.'
        }), 202

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@report_bp.route('/api/relatorio_higienizacao/validate_vtadmin', methods=['POST'])
@login_required
def validate_relatorio_higienizacao_vtadmin():
    if not user_can_manage_card_hygiene():
        return jsonify({'error': 'Você não tem permissão para executar higienização de cadastro.'}), 403

    data = request.json or {}
    vtadmin_username = str(data.get('vtadmin_username') or '').strip()
    vtadmin_password = data.get('vtadmin_password') or ''

    is_valid, message = validate_vtadmin_credentials(vtadmin_username, vtadmin_password)
    if not is_valid:
        return jsonify({'ok': False, 'error': message}), 400

    return jsonify({'ok': True, 'message': message}), 200


@report_bp.route('/api/relatorio_higienizacao/cancel', methods=['POST'])
@login_required
def cancel_relatorio_higienizacao():
    if not user_can_manage_card_hygiene():
        return jsonify({'error': 'Você não tem permissão para cancelar higienização de cadastro.'}), 403

    data = request.json or {}
    job_id = (data.get('job_id') or data.get('cancel_token') or '').strip()
    if not job_id:
        return jsonify({'error': 'job_id não informado.'}), 400

    # Only the job owner or a monitor-perm user can cancel
    can_monitor = user_has_permission('acompanhar_higienizacao')
    if not can_monitor:
        job = get_job_from_db(job_id)
        if job and job.get('user_id') != session.get('user_id'):
            return jsonify({'error': 'Você não pode cancelar a higienização de outro usuário.'}), 403

    request_cancel_job(job_id)
    return jsonify({'ok': True, 'cancel_requested': True}), 200


@report_bp.route('/api/relatorio_higienizacao/status', methods=['GET'])
@login_required
def status_relatorio_higienizacao():
    job_id = (request.args.get('job_id') or request.args.get('cancel_token') or '').strip()
    if not job_id:
        return jsonify({'error': 'job_id não informado.'}), 400

    job = get_job_from_db(job_id)
    if not job:
        return jsonify({'ok': False, 'found': False}), 200

    # Only owner or monitor-perm users can query status
    can_monitor = user_has_permission('acompanhar_higienizacao')
    if not can_monitor and job.get('user_id') != session.get('user_id'):
        return jsonify({'error': 'Acesso negado.'}), 403

    result_data = None
    if job.get('result_json'):
        try:
            result_data = json.loads(job['result_json'])
        except Exception:
            pass

    return jsonify({
        'ok': True,
        'found': True,
        'job_id': job.get('id'),
        'status': job.get('status'),
        'percent': job.get('progress_percent', 0),
        'label': job.get('progress_label', ''),
        'detail': job.get('progress_detail', ''),
        'current_card': 'Protegido' if job.get('current_card') else None,
        'current_cpf': 'Protegido' if job.get('current_cpf') else None,
        'processed': job.get('processed', 0),
        'total': job.get('total', 0),
        'username': job.get('username'),
        'started_at': job.get('started_at'),
        'finished_at': job.get('finished_at'),
        'result': result_data,
    }), 200


@report_bp.route('/api/relatorio_higienizacao/export_result', methods=['GET'])
@login_required
def export_relatorio_higienizacao_result():
    job_id = (request.args.get('job_id') or '').strip()
    if not job_id:
        return jsonify({'error': 'job_id não informado.'}), 400

    job = get_job_from_db(job_id)
    if not job:
        return jsonify({'error': 'Lote não encontrado.'}), 404

    can_monitor = user_has_permission('acompanhar_higienizacao')
    if not can_monitor and job.get('user_id') != session.get('user_id'):
        return jsonify({'error': 'Acesso negado.'}), 403

    result_data = {}
    if job.get('result_json'):
        try:
            result_data = json.loads(job['result_json']) or {}
        except Exception:
            result_data = {}

    success_items = result_data.get('success_items') or []
    failed_items = result_data.get('failed_items') or []
    skipped_items = result_data.get('skipped_items') or []

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = 'Resumo'
        ws_success = wb.create_sheet('Sucessos')
        ws_failed = wb.create_sheet('Falhas')
        ws_skipped = wb.create_sheet('Ignorados')

        title_fill = PatternFill(start_color='062B3D', end_color='062B3D', fill_type='solid')
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        section_fill = PatternFill(start_color='0F3B54', end_color='0F3B54', fill_type='solid')
        alt_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        white_font = Font(color='FFFFFF', bold=True)
        title_font = Font(color='67E8F9', bold=True, size=13)
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        ws_summary.merge_cells('A1:B1')
        ws_summary['A1'] = 'Relatório de Higienização de Cadastro'
        ws_summary['A1'].font = title_font
        ws_summary['A1'].fill = title_fill
        ws_summary['A1'].alignment = Alignment(horizontal='center')

        summary_rows = [
            ('ID do lote', job.get('id') or ''),
            ('Usuário DataCross', job.get('username') or ''),
            ('Status', job.get('status') or ''),
            ('Iniciado em', job.get('started_at') or ''),
            ('Finalizado em', job.get('finished_at') or ''),
            ('Observação', job.get('observation') or ''),
            ('Total solicitado', int(job.get('total') or 0)),
            ('Processados', int(job.get('processed') or 0)),
            ('Sucessos', int(result_data.get('success_count') or len(success_items))),
            ('Falhas', int(result_data.get('failure_count') or len(failed_items))),
            ('Ignorados', int(result_data.get('skip_count') or len(skipped_items))),
        ]

        ws_summary['A3'] = 'Resumo do Lote'
        ws_summary['A3'].font = white_font
        ws_summary['A3'].fill = section_fill

        for idx, (label, value) in enumerate(summary_rows, start=4):
            ws_summary[f'A{idx}'] = label
            ws_summary[f'B{idx}'] = value
            ws_summary[f'A{idx}'].border = thin_border
            ws_summary[f'B{idx}'].border = thin_border
            if idx % 2 == 0:
                ws_summary[f'A{idx}'].fill = alt_fill
                ws_summary[f'B{idx}'].fill = alt_fill

        success_headers = ['CPF', 'Nome', 'Cartão', 'Data Nascimento', 'Origem Data Nascimento', 'Observação Automática']
        for col_idx, header in enumerate(success_headers, start=1):
            cell = ws_success.cell(row=1, column=col_idx, value=header)
            cell.font = white_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for row_idx, item in enumerate(success_items, start=2):
            values = [
                'Protegido',
                'Protegido',
                'Protegido',
                'Protegido',
                item.get('birthdate_source', ''),
                item.get('note', ''),
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws_success.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        failed_headers = ['CPF', 'Nome', 'Cartão', 'Data Nascimento', 'Origem Data Nascimento', 'Observação Automática', 'Erro']
        for col_idx, header in enumerate(failed_headers, start=1):
            cell = ws_failed.cell(row=1, column=col_idx, value=header)
            cell.font = white_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for row_idx, item in enumerate(failed_items, start=2):
            values = [
                'Protegido',
                'Protegido',
                'Protegido',
                'Protegido',
                item.get('birthdate_source', ''),
                item.get('note', ''),
                item.get('error', ''),
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws_failed.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        skipped_headers = ['CPF', 'Nome', 'Cartão', 'Data Nascimento', 'Origem Data Nascimento', 'Observação Automática', 'Status', 'Motivo']
        for col_idx, header in enumerate(skipped_headers, start=1):
            cell = ws_skipped.cell(row=1, column=col_idx, value=header)
            cell.font = white_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for row_idx, item in enumerate(skipped_items, start=2):
            values = [
                'Protegido',
                'Protegido',
                'Protegido',
                'Protegido',
                item.get('birthdate_source', ''),
                item.get('note', ''),
                item.get('status', ''),
                item.get('reason', ''),
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws_skipped.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        for ws in (ws_summary, ws_success, ws_failed, ws_skipped):
            for col_idx in range(1, ws.max_column + 1):
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(col_idx)
                for row_idx in range(1, ws.max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    cell_value = '' if cell_value is None else str(cell_value)
                    max_length = max(max_length, len(cell_value))
                ws.column_dimensions[column_letter].width = min(max_length + 3, 60)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M')
        filename = f'higienizacao_lote_{job_id}_{timestamp}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )
    except Exception as exc:
        return jsonify({'error': f'Falha ao exportar relatório: {exc}'}), 500


@report_bp.route('/api/relatorio_higienizacao/jobs/mine', methods=['GET'])
@login_required
def my_hygiene_jobs():
    """Returns visible hygiene jobs for the current user."""
    jobs = get_active_jobs_for_user(session['user_id'])
    return jsonify({'jobs': jobs}), 200


@report_bp.route('/api/relatorio_higienizacao/popup/close', methods=['POST'])
@login_required
def close_hygiene_popup():
    data = request.json or {}
    job_id = (data.get('job_id') or '').strip()
    if not job_id:
        return jsonify({'error': 'job_id nÃ£o informado.'}), 400

    can_monitor = user_has_permission('acompanhar_higienizacao')
    closed = set_hygiene_job_popup_closed(job_id, session['user_id'], can_monitor=can_monitor)
    if not closed:
        return jsonify({'error': 'Lote nÃ£o encontrado para este usuÃ¡rio.'}), 404
    return jsonify({'ok': True}), 200


@report_bp.route('/api/relatorio_higienizacao/jobs/monitor', methods=['GET'])
@login_required
def monitor_hygiene_jobs():
    """Returns all currently active hygiene jobs (requires perm_acompanhar_higienizacao)."""
    can_monitor = user_has_permission('acompanhar_higienizacao')
    if not can_monitor:
        return jsonify({'error': 'Você não tem permissão para monitorar higienizações de outros usuários.'}), 403
    jobs = get_all_active_jobs()
    return jsonify({'jobs': jobs}), 200


def _detect_cpf_column(df):
    """Detecta a coluna CPF por nome ou por conteúdo."""
    cpf_names = {'cpf', 'nr_cpf', 'num_cpf', 'numero_cpf', 'cpf_cliente',
                 'cpfcliente', 'documento', 'doc', 'cpf_aluno', 'cpf_beneficiario',
                 'cpfbeneficiario', 'nr_documento', 'num_documento'}
    for col in df.columns:
        if col.strip().lower().replace(' ', '_') in cpf_names:
            return col
    # Fallback: coluna com maior número de valores que parecem CPF (11 dígitos)
    best_col, best_count = df.columns[0], 0
    sample = df.head(200)
    for col in df.columns:
        count = int(sample[col].apply(
            lambda x: len(re.sub(r'\D', '', str(x))) == 11
        ).sum())
        if count > best_count:
            best_count, best_col = count, col
    return best_col


@report_bp.route('/api/preview_planilha', methods=['POST'])
@login_required
@permission_required('cruzamento')
def preview_planilha():
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'Arquivo não enviado'}), 400
    try:
        fname = (file.filename or '').lower()
        if fname.endswith('.csv'):
            df_full = pd.read_csv(file, sep=None, engine='python', dtype=str, encoding='latin1')
        elif fname.endswith('.xlsx') or fname.endswith('.xls'):
            df_full = pd.read_excel(file, dtype=str)
        else:
            return jsonify({'error': 'Formato não suportado'}), 400

        total_rows = len(df_full)
        df = df_full.head(500)

        cpf_col = _detect_cpf_column(df)

        def _is_valid_cpf(x):
            d = re.sub(r'\D', '', str(x))
            return len(d) == 11 or len(d) == 10  # 10 = zero à esquerda perdido pelo Excel

        cpfs_validos = int(df_full[cpf_col].apply(_is_valid_cpf).sum()) if cpf_col else 0

        outras = [c for c in df.columns if c != cpf_col]
        amostra = df[[cpf_col] + outras[:4]].head(4).fillna('').to_dict(orient='records')

        return jsonify({
            'colunas': list(df.columns),
            'cpf_coluna': cpf_col,
            'cpfs_validos': cpfs_validos,
            'amostra_lida': len(df),
            'total_rows': total_rows,
            'outras_colunas': outras,
            'amostra': amostra,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@report_bp.route('/download_template')
@login_required
def download_template():
    output = io.BytesIO()
    df = pd.DataFrame({
        'CPF': ['705.469.112-18', '002.218.042-77', '123.456.789-09',
                '111.222.333-44', '555.666.777-88'],
        'NOME_OPCIONAL': ['João da Silva', 'Maria Santos', 'Pedro Lima',
                          'Ana Costa', 'Carlos Souza'],
    })
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Dados')
        wb = writer.book
        ws = writer.sheets['Dados']
        hdr = wb.add_format({'bold': True, 'bg_color': '#1E293B', 'font_color': 'white', 'border': 1})
        for i, col in enumerate(df.columns):
            ws.write(0, i, col, hdr)
            ws.set_column(i, i, 22)

        inst = wb.add_worksheet('Instruções')
        bold = wb.add_format({'bold': True})
        inst.write(0, 0, 'Como usar o Cruzamento em Massa', bold)
        instrucoes = [
            '1. A primeira coluna DEVE conter os CPFs (com ou sem pontos e traços)',
            '2. Você pode adicionar outras colunas — elas serão mantidas no resultado',
            '3. Formatos aceitos para CPF: 000.000.000-00  ou  00000000000',
            '4. Salve como .xlsx ou .csv antes de importar',
            '5. Tamanho máximo: 50 MB',
        ]
        for i, linha in enumerate(instrucoes, start=2):
            inst.write(i, 0, linha)
        inst.set_column(0, 0, 65)

    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name='MODELO_CRUZAMENTO.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@report_bp.route('/upload_file', methods=['POST'])
@login_required
@permission_required('cruzamento')
def upload():
    cleanup_uploads_folder(3)

    if 'file' not in request.files:
        return jsonify({'error': 'Arquivo não enviado'}), 400
    file = request.files['file']
    original_filename = secure_filename(file.filename or '')
    if not original_filename.lower().endswith(('.csv', '.xlsx')):
        return jsonify({'error': 'Apenas arquivos .csv ou .xlsx sao permitidos'}), 400
    file_id = str(uuid.uuid4())[:8]
    unique_filename = f"{file_id}_{original_filename}"
    if not os.path.exists(UPLOAD_FOLDER):
        os.makedirs(UPLOAD_FOLDER)
    path = os.path.join(UPLOAD_FOLDER, unique_filename)
    file.save(path)
    return jsonify({'filename': unique_filename}), 200


@report_bp.route('/download/<path:filename>')
@login_required
def download(filename):
    if not (user_has_permission('relatorio') or user_has_permission('cruzamento')):
        return "Acesso negado", 403
    file_path = safe_join(UPLOAD_FOLDER, filename)
    if not file_path:
        return "File not found", 404
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    return "File not found", 404
